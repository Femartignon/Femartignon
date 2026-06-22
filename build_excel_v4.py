"""
Fase 3 — Automação
  - CF visual alerts em todos os DB_ sheets
  - Nova coluna Prazo em DB_ATIVIDADES
  - Aba ⚡ ALERTAS com fórmulas calculadas
  - Roadmap Fase 3 = concluída
"""
import re, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from collections import defaultdict, Counter

# ── Styles ────────────────────────────────────────────────────────────────────
def fx(h): return PatternFill("solid", fgColor=h)
def ft(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def bd():
    s=Side(style="thin",color="D9D9D9"); return Border(left=s,right=s,top=s,bottom=s)
def bdM(col="2E75B6"):
    s=Side(style="medium",color=col); return Border(left=s,right=s,top=s,bottom=s)
def al(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def set_w(ws,c,w): ws.column_dimensions[get_column_letter(c)].width=w

C_NAVY,C_BLUE,C_LBLUE = "1F3864","2E75B6","BDD7EE"
C_LGRAY,C_WHITE       = "F2F2F2","FFFFFF"
C_GREEN               = "375623"
C_ORANGE              = "843C0C"
C_RED                 = "C00000"
C_YELLOW              = "7F6000"
C_PURPLE              = "7030A0"
C_TEAL                = "1F6B75"
C_LRED,C_LORANGE,C_LYELL,C_LGREEN,C_LBLUE2 = "FFDDD8","FCE4D6","FFF2CC","E2EFDA","BDD7EE"

STATUS_FX = {
    'Realizado':('E2EFDA','375623'),'a Realizar':('BDD7EE','1F3864'),
    'Cancelado':('FFDDD8','C00000'),'CONCLUÍDO':('E2EFDA','375623'),
    'EM ANDAMENTO':('FFF2CC','7F6000'),'PENDENTE':('FCE4D6','843C0C'),
    'ATRASADO':('FFDDD8','C00000'),'-':('F2F2F2','404040'),
    'Pago':('E2EFDA','375623'),'Pós Evento':('FFF2CC','7F6000'),
    'Pendente':('FCE4D6','843C0C'),'aprovado':('BDD7EE','1F3864'),
    'pago':('E2EFDA','375623'),'Recebida':('E2EFDA','375623'),
    'Recebido':('E2EFDA','375623'),'Homologado':('E2EFDA','375623'),
    'A preencher':('FFF2CC','7F6000'),
}
def scell(cell,status):
    k=status.strip() if status else ''
    if k in STATUS_FX:
        bg,fg=STATUS_FX[k]; cell.fill=fx(bg); cell.font=ft(True,fg,9)
    cell.alignment=al(); cell.border=bd()

def hdr(ws,r,cols,bg=C_NAVY,fg="FFFFFF",sz=10):
    for c in cols:
        cell=ws.cell(row=r,column=c)
        cell.fill=fx(bg); cell.font=ft(True,fg,sz); cell.alignment=al(); cell.border=bd()

def drow(ws,r,mc,alt=False):
    bg=C_LGRAY if alt else C_WHITE
    for c in range(1,mc+1):
        cell=ws.cell(row=r,column=c)
        cell.fill=fx(bg); cell.font=ft(size=9); cell.alignment=al("left"); cell.border=bd()

def tblock(ws,mr,text,bg,fg="FFFFFF",sz=14):
    ws.merge_cells(mr)
    fc=mr.split(':')[0]; c=ws[fc]; c.value=text
    c.fill=fx(bg); c.font=ft(True,fg,sz); c.alignment=al()

def sblock(ws,mr,text,bg,fg):
    ws.merge_cells(mr)
    fc=mr.split(':')[0]; c=ws[fc]; c.value=text
    c.fill=fx(bg); c.font=ft(False,fg,9); c.alignment=al("left")

def adv(ws,f1,sqref,pt="",pr=""):
    dv=DataValidation(type="list",formula1=f1,showDropDown=False,
        showInputMessage=bool(pt),showErrorMessage=True)
    if pt: dv.promptTitle=pt
    if pr: dv.prompt=pr
    dv.sqref=sqref; ws.add_data_validation(dv)

def add_cf(ws,rng,formula,bg,fg,stop=False):
    """Add FormulaRule conditional formatting."""
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[formula],
        fill=fx(bg),
        font=ft(True,fg,9),
        stopIfTrue=stop
    ))

TIPO_MAP={'C':'Cota','A':'Apoio','S':'Simpósio','T':'TOME/Advisory'}

# ── Load source ───────────────────────────────────────────────────────────────
src=load_workbook('/root/.claude/uploads/6d41cea8-3482-52e3-859c-cdd7002a8a66/1318286f-Calenda_rio_Neuro_Acesso_NEC.xlsx',data_only=True)

def clean_em(v):
    if not v: return ''
    return re.sub(r'\s+','',str(v).strip().upper())
def fix_name(v):
    fixes={'Biansa Fasas':'Bianca Facas','biansa fasas':'Bianca Facas'}
    s=str(v).strip() if v else ''
    return fixes.get(s,s)

# ── Extract events ────────────────────────────────────────────────────────────
cal=src['Calendário de Eventos']
events=[]; em_set=set()
for row in cal.iter_rows(min_row=4,max_row=60,values_only=True):
    if row[2] is None: continue
    name=str(row[2]).strip()
    if name in ('APOIO','COTA','SIMPÓSIO','TOME','','-'): continue
    def s(v): return str(v).strip() if v is not None else ''
    def d(v):
        if isinstance(v,datetime.datetime): return v.date()
        if isinstance(v,datetime.date): return v
        return None
    try: psp=int(float(str(row[19]))) if row[19] not in (None,'','-','X') else None
    except: psp=None
    try: psa=int(float(str(row[20]))) if row[20] not in (None,'','-','X') else None
    except: psa=None
    em_raw=s(row[4]); em_clean=clean_em(em_raw)
    sol=fix_name(s(row[9]))
    dt_ini=d(row[13]); dt_fim=d(row[14])
    q=f"Q{(dt_ini.month-1)//3+1} {s(row[12])}" if dt_ini and s(row[12]) else ''
    dup='⚠️ EM dup.' if em_clean and em_clean in em_set and em_clean not in ('PENDENTE','TBD','') else ''
    if em_clean and em_clean not in ('PENDENTE','TBD',''): em_set.add(em_clean)
    events.append({'ID_Unico':em_clean or f"SEM-EM-{len(events)+1:03d}",
        'Evento':name,'Tipo':s(row[1]),'EM':em_raw,'EM_P':em_clean,
        'StatusEv':s(row[7]),'StatusEM':s(row[8]),'Sol':sol,'BU':s(row[10]),
        'Produto':s(row[11]),'Ano':s(row[12]),'DtIni':dt_ini,'DtFim':dt_fim,'Q':q,
        'Cidade':s(row[16]),'Formato':s(row[17]),'Local':s(row[18]),
        'PSP':psp,'PSA':psa,'Agencia':s(row[21]),'Hotel':s(row[22]),
        'Sociedade':s(row[23]),'Stand':s(row[24]),'Comp':s(row[3]),
        'Obs':s(row[27]),'Flag':dup})

# ── Extract activities ────────────────────────────────────────────────────────
ACT_CFG={
    'Fórum de Acesso 2026':('FÓRUM ACESSO 2026',1,2,3,4,5,6,5),
    ' FÓRUM GEDIIB ':('FORUM GEDIIB BRASÍLIA',3,4,5,6,7,8,5),
    'Cong. Medicina do Sono':('Congresso Medicina do Sono',1,2,3,4,5,6,5),
    'Brain 2026':('BRAIN 2026',2,3,4,5,6,7,6),
    'Brain Simpósio':('BRAIN 2026 - Simpósio',2,3,4,5,6,7,5),
    'Board Albert Einstein':('Board Review Albert Einstein',1,2,3,4,5,6,5),
    'CONASENS':('CONASENS',1,2,3,4,5,6,5),
    'Aprender Criança 2026':('APRENDER CRIANÇA 2026',1,2,3,4,5,6,5),
    'ADV. BOARD FRUZAQLA':('Advisory Board Fruzaqla',1,2,3,4,5,6,5),
    'Masterclass 2026':('MASTERCLASS 2026',2,3,4,5,6,7,5),
    'Masterclass Simpósio':('MASTERCLASS 2026 - Simpósio',2,3,4,5,6,7,5),
    'ABENEPI':('ABENEPI 2026',1,2,3,4,5,6,6),
}
activities=[]
for sname,(ev,ce,ca,cr,cp,cs,co,start) in ACT_CFG.items():
    ws2=src[sname]; last_e=''
    for row in ws2.iter_rows(min_row=start,max_row=ws2.max_row,values_only=True):
        raw=list(row)+[None]*15
        ev_v=raw[ce-1];av=raw[ca-1];rv=raw[cr-1];pv=raw[cp-1];sv=raw[cs-1];ov=raw[co-1]
        if ev_v not in (None,'') and str(ev_v).strip() not in ('Atividade','Responsável','Andamento (%)','Status','Observações'):
            last_e=str(ev_v).strip()
        if av is None or str(av).strip() in ('','Atividade','Responsável','Andamento (%)','Status','Observações'): continue
        try: pct=int(float(str(pv))) if pv is not None else 0
        except: pct=0
        stat=str(sv).strip() if sv else ''
        if not stat: stat='CONCLUÍDO' if pct==100 else ('EM ANDAMENTO' if pct>0 else 'PENDENTE')
        activities.append({'Evento':ev,'Etapa':last_e,'Ativ':str(av).strip(),
            'Resp':str(rv).strip() if rv else '','Prazo':'A definir',
            'Pct':pct,'Status':stat,'Obs':str(ov).strip() if ov else ''})

# ── Extract financeiro ────────────────────────────────────────────────────────
fin_ws=src['FINANCEIRO']
financeiro=[]
for row in fin_ws.iter_rows(min_row=2,max_row=fin_ws.max_row,values_only=True):
    if row[0] is None: continue
    def n(v):
        if v is None: return 0.0
        try: return float(str(v).replace(',','.'))
        except: return 0.0
    financeiro.append({'Evento':str(row[0]).strip(),'EM':clean_em(row[1]),
        'FY':str(row[2]).strip() if row[2] else '',
        'Forn':str(row[3]).strip() if row[3] else '',
        'Tipo':str(row[4]).strip() if row[4] else '',
        'PO':str(row[5]).strip() if row[5] else '',
        'Prev':n(row[6]),'R25':n(row[7]),'R26':n(row[8]),'Tot':n(row[9]),
        'StatPag':str(row[10]).strip() if row[10] else '',
        'StatNF':str(row[11]).strip() if row[11] else '',
        'Obs':str(row[12]).strip() if row[12] else ''})

SUPPLIER_CATALOG={
    'Casa da Vila':('Agência de Produção','Produção de Eventos','SP','Homologado'),
    'Incentivare':('Agência Logística','Logística, Buffet, Apoios','SP','Homologado'),
    'OPUS':('Agência Logística','Logística Convidados','SP','Homologado'),
    'Opus':('Agência Logística','Logística Convidados','SP','Homologado'),
    'Jardim Elétrico':('Montadora de Stand','Montagem de Stands','SP','Homologado'),
    'Jardim':('Montadora de Stand','Montagem de Stands','SP','A verificar'),
    'Poli Design':('Montadora de Stand','Montagem de Stands','SP','Homologado'),
    'Poli':('Montadora de Stand','Montagem de Stands','SP','A verificar'),
    'Boutique Gourmet':('Buffet/Alimentação','Buffet, Lunch Box','SP','Homologado'),
    'Boutique':('Buffet/Alimentação','Buffet','SP','A verificar'),
    'APM':('Sociedade Médica','Cota de Patrocínio','SP','Contrato'),
    'Sociedade':('Sociedade Médica','Cota de Patrocínio','Vários','Contrato'),
    'Dr. Fernando':('Palestrante','Honorários Palestrante','SP','Contrato'),
    'Dra. Dalva':('Palestrante','Honorários Palestrante','SP','Contrato'),
    'FEE Palestrante':('Palestrante','Honorários Palestrante','Vários','A verificar'),
    'Oficial':('Montadora Oficial','Montagem Oficial do Evento','Vários','Oficial evento'),
    'Homologado local':('Buffet Local','Buffet homologado pelo local','Vários','Homologado'),
}
RESP_UNICOS=sorted(set(a['Resp'] for a in activities if a['Resp']))
BU_LIST=sorted(set(e['BU'] for e in events if e['BU']))
STATUS_EV=['Realizado','a Realizar','Cancelado','Em Planejamento']
STATUS_EM=['aprovada','Encerrada','pendente','solicitado','Sem ação Eventos']
STATUS_AT=['CONCLUÍDO','EM ANDAMENTO','PENDENTE','ATRASADO','-']
STATUS_PAG=['Pago','pago','aprovado','Pós Evento','Pendente','A preencher','Cancelado']
STATUS_NF=['Recebida','Recebido','Adiantamento','Pendente','A preencher']
TIPO_EV=['Cota','Apoio','Simpósio','TOME/Advisory']
FORMATO=['Presencial','Virtual','Hibrido']
COMPLEXIDADE=['Alta','Média','Baixa']
AGENCIA=['Incentivare','OPUS','Opus','A preencher']
TIPO_CUSTO=['Cota de Patrocínio','Logística Convidados','Produção','Stand',
    'Buffet Stand','Lunch Box','Palestrante','FEE Palestrante','Agência','Outros']
STATUS_FORN=['Homologado','A verificar','Contrato','Oficial evento','A preencher','Suspenso']

print(f"Eventos:{len(events)} | Atividades:{len(activities)} | Financeiro:{len(financeiro)}")

# ═══════════════════════════════════════════════════════════════════════════════
# Build workbook
# ═══════════════════════════════════════════════════════════════════════════════
wb=Workbook(); wb.remove(wb.active)
n_ev=len(events); n_at=len(activities); n_fin=len(financeiro)

# ═══════════════════════════════════════════════════════════════════════════════
# INÍCIO
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("🏠 INÍCIO")
ws.sheet_view.showGridLines=False
for c,w in zip(range(1,9),[2,30,28,28,22,22,22,8]):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.row_dimensions[1].height=10
for r in range(2,8): ws.row_dimensions[r].height=24
for r in range(2,8):
    for col in range(2,9): ws.cell(row=r,column=col).fill=fx(C_NAVY)
ws.merge_cells('B2:H4'); c=ws['B2']; c.value="GESTÃO DE EVENTOS FARMACÊUTICOS"
c.font=ft(True,"FFFFFF",22); c.alignment=al()
ws.merge_cells('B5:H5'); c=ws['B5']; c.value="Neurociências & Acesso  ·  NEC  ·  Takeda"
c.font=ft(False,"BDD7EE",13); c.alignment=al()
ws.merge_cells('B6:H7'); c=ws['B6']
c.value=f"Versão 4  ·  Fase 3 Concluída  ·  {datetime.date.today().strftime('%d/%m/%Y')}"
c.font=ft(False,"9DC3E6",10); c.alignment=al()
ws.row_dimensions[9].height=55; ws.row_dimensions[10].height=22
navs=[
    ('B','DB_EVENTOS',"📋 Banco de\nEventos",C_BLUE,f"{n_ev} eventos"),
    ('C','DB_ATIVIDADES',"✅ Banco de\nAtividades","375623",f"{n_at} atividades + Prazo"),
    ('D','DB_FINANCEIRO',"💰 Banco\nFinanceiro","843C0C",f"{n_fin} lançamentos"),
    ('E','DB_FORNECEDORES',"🏢 Banco de\nFornecedores",C_TEAL,f"{len(SUPPLIER_CATALOG)} cadastros"),
    ('F','DB_PARTICIPANTES',"👥 Participantes",C_PURPLE,f"25 pessoas"),
    ('G','⚡ ALERTAS',"⚡ Alertas\nVisuais","C4700E","Tempo real"),
]
for col_l,_,label,bg,sub in navs:
    col=ord(col_l)-ord('A')+1
    c=ws.cell(row=9,column=col); c.value=label
    c.fill=fx(bg); c.font=ft(True,"FFFFFF",11); c.alignment=al(); c.border=bdM()
    c=ws.cell(row=10,column=col); c.value=sub
    c.fill=fx(C_LGRAY); c.font=ft(False,"404040",8); c.alignment=al(); c.border=bd()
ws.row_dimensions[12].height=18; ws.merge_cells('B12:H12')
total_prev=sum(f['Prev'] for f in financeiro)
c=ws['B12']; c.value=(f"  Budget Total: R$ {total_prev:,.0f}  ·  A Realizar: "
    f"{sum(1 for e in events if 'Realizar' in e['StatusEv'])}  ·  "
    f"Pendentes: {sum(1 for a in activities if a['Status']=='PENDENTE')}  ·  "
    f"⚡ Fase 3 - Alertas Visuais Ativos  ·  Scripts M365 prontos")
c.fill=fx(C_LBLUE); c.font=ft(True,C_NAVY,10); c.alignment=al("left")
ws.row_dimensions[14].height=16; ws.merge_cells('B14:H14')
ws['B14'].value=("  ⚠️  Abas com CF ativa: DB_EVENTOS, DB_ATIVIDADES, DB_FINANCEIRO  ·  "
    "Office Script: automation/office_script_nec.ts  ·  Power Query: automation/power_query_nec.pq")
ws['B14'].fill=fx(C_LYELL); ws['B14'].font=ft(False,C_YELLOW,9); ws['B14'].alignment=al("left")

# ═══════════════════════════════════════════════════════════════════════════════
# DB_EVENTOS + CF rules
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("DB_EVENTOS")
ws.sheet_view.showGridLines=False; ws.freeze_panes="D4"
tblock(ws,'A1:Z1',"📋 BANCO DE DADOS CENTRAL DE EVENTOS  ·  v4  ·  Alertas Visuais Ativos",C_NAVY,sz=12)
ws.row_dimensions[1].height=28
sblock(ws,'A2:Z2',
    "🔴 Linha vermelha = evento vencido (a Realizar + data passada)  ·  "
    "🟠 Laranja = próximos 7 dias  ·  🟡 Amarelo = próximos 30 dias",C_LBLUE,C_NAVY)
ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=32

COLS_EV=[
    (1,'ID','A',5),(2,'ID_Unico','B',16),(3,'Flag','C',10),(4,'Evento','D',40),
    (5,'Tipo','E',12),(6,'EM Original','F',16),(7,'Status Evento','G',14),(8,'Status EM','H',14),
    (9,'Ano','I',6),(10,'Trimestre','J',10),(11,'Data Início','K',12),(12,'Data Fim','L',12),
    (13,'BU','M',16),(14,'Produto','N',18),(15,'Solicitante','O',18),(16,'Cidade','P',16),
    (17,'Formato','Q',12),(18,'Complexidade','R',12),(19,'Local','S',28),
    (20,'Agência','T',16),(21,'Hotel','U',20),(22,'Sociedade','V',30),
    (23,'Stand','W',10),(24,'PSs Plan.','X',9),(25,'PSs Pres.','Y',9),(26,'Observações','Z',38),
]
for ci,label,col_l,w in COLS_EV: ws.column_dimensions[get_column_letter(ci)].width=w
hdr(ws,3,range(1,27))
for ci,label,_,_ in COLS_EV: ws.cell(row=3,column=ci).value=label

tab=Table(displayName="TblEventos",ref=f"A3:Z{3+n_ev}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tab)

for i,ev in enumerate(events):
    r=4+i; drow(ws,r,26,alt=(i%2==1))
    row_data=[i+1,ev['ID_Unico'],ev['Flag'],ev['Evento'],
        TIPO_MAP.get(ev['Tipo'],ev['Tipo']),ev['EM'],ev['StatusEv'],ev['StatusEM'],
        ev['Ano'],ev['Q'],ev['DtIni'],ev['DtFim'],
        ev['BU'],ev['Produto'],ev['Sol'],ev['Cidade'],
        ev['Formato'],ev['Comp'],ev['Local'],
        ev['Agencia'],ev['Hotel'],ev['Sociedade'],
        ev['Stand'],ev['PSP'],ev['PSA'],ev['Obs']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (11,12) and isinstance(val,datetime.date):
            cell.number_format='DD/MM/YYYY'
        if c_idx==7: scell(cell,ev['StatusEv'])
        if c_idx==8: scell(cell,ev['StatusEM'])
        if c_idx==3 and val:
            cell.fill=fx(C_LRED); cell.font=ft(True,C_RED,9); cell.alignment=al()

# ── Conditional Formatting rules — DB_EVENTOS ─────────────────────────────────
cf_rng=f"A4:Z{3+n_ev}"
# Priority 1 (highest): Overdue events — red
add_cf(ws,cf_rng,'=AND($G4="a Realizar",$K4<TODAY(),$K4<>"",ISNUMBER($K4))',C_LRED,C_RED,stop=False)
# Priority 2: Upcoming 7 days — orange
add_cf(ws,cf_rng,'=AND($G4="a Realizar",ISNUMBER($K4),$K4>=TODAY(),$K4<=TODAY()+7)',C_LORANGE,C_ORANGE)
# Priority 3: Upcoming 30 days — yellow
add_cf(ws,cf_rng,'=AND($G4="a Realizar",ISNUMBER($K4),$K4>TODAY()+7,$K4<=TODAY()+30)',C_LYELL,C_YELLOW)
# Priority 4: Realized — subtle green
add_cf(ws,cf_rng,'=$G4="Realizado"',"F0FFF0","375623")

# Validations
adv(ws,f'"{",".join(STATUS_EV)}"',f"G4:G{3+n_ev}","Status do Evento")
adv(ws,f'"{",".join(STATUS_EM)}"',f"H4:H{3+n_ev}","Status EM")
adv(ws,f'"{",".join(TIPO_EV)}"',f"E4:E{3+n_ev}","Tipo")
adv(ws,f'"{",".join(BU_LIST)}"',f"M4:M{3+n_ev}","Business Unit")
adv(ws,f'"{",".join(FORMATO)}"',f"Q4:Q{3+n_ev}","Formato")
adv(ws,f'"{",".join(COMPLEXIDADE)}"',f"R4:R{3+n_ev}","Complexidade")
adv(ws,f'"{",".join(AGENCIA)}"',f"T4:T{3+n_ev}","Agência")

print("DB_EVENTOS ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# DB_ATIVIDADES + Prazo column + CF
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("DB_ATIVIDADES")
ws.sheet_view.showGridLines=False; ws.freeze_panes="D4"
tblock(ws,'A1:I1',"✅ BANCO DE DADOS DE ATIVIDADES  ·  v4  ·  Coluna Prazo + Alertas CF","375623",sz=12)
ws.row_dimensions[1].height=28
sblock(ws,'A2:I2',
    "🔴 Prazo vencido + PENDENTE = ATRASADO  ·  🟠 Prazo em 7 dias  ·  "
    "🟡 PENDENTE geral  ·  🟢 CONCLUÍDO  ·  Coluna Prazo: insira data DD/MM/AAAA",
    "E2EFDA",C_GREEN)
ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=32

# NEW: Prazo column added as col F (status → H, % → G)
COLS_AT=[
    (1,'ID','A',5),(2,'Evento','B',38),(3,'Etapa','C',22),(4,'Atividade','D',44),
    (5,'Responsável','E',18),(6,'Prazo','F',13),(7,'% Conclusão','G',13),
    (8,'Status','H',16),(9,'Observações','I',40),
]
for ci,label,col_l,w in COLS_AT: ws.column_dimensions[get_column_letter(ci)].width=w
hdr(ws,3,range(1,10),bg="375623")
for ci,label,_,_ in COLS_AT: ws.cell(row=3,column=ci).value=label

tab2=Table(displayName="TblAtividades",ref=f"A3:I{3+n_at}")
tab2.tableStyleInfo=TableStyleInfo(name="TableStyleMedium7",showRowStripes=True)
ws.add_table(tab2)

for i,at in enumerate(activities):
    r=4+i; drow(ws,r,9,alt=(i%2==1))
    pct=at['Pct']
    row_data=[i+1,at['Evento'],at['Etapa'],at['Ativ'],at['Resp'],
              None,pct/100 if pct else 0,at['Status'],at['Obs']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx==7:
            cell.number_format='0%'; cell.alignment=al()
            if pct==100: cell.fill=fx("E2EFDA"); cell.font=ft(True,"375623",9)
            elif pct>=50: cell.fill=fx("FFF2CC"); cell.font=ft(True,"7F6000",9)
            elif pct>0: cell.fill=fx("FCE4D6"); cell.font=ft(True,"843C0C",9)
            else: cell.fill=fx("FFDDD8"); cell.font=ft(True,"C00000",9)
        if c_idx==8: scell(cell,at['Status'])
        if c_idx==6:
            # Prazo: célula de DATA (vazia, pronta para receber DD/MM/AAAA)
            cell.number_format='DD/MM/YYYY'
            cell.font=ft(False,"7F6000",8); cell.fill=fx("FFFFF0"); cell.alignment=al("center")

# ── CF rules — DB_ATIVIDADES ─────────────────────────────────────────────────
# DataBar adicionado DEPOIS das FormulaRules para não interferir com prioridades
cf_rng_at=f"A4:I{3+n_at}"
# P1 (prioridade 1, mais alta): ATRASADO → RED (status explícito)
add_cf(ws,cf_rng_at,'=$H4="ATRASADO"',C_LRED,C_RED)
# P2: Prazo vencido + PENDENTE → RED (vencimento implícito)
add_cf(ws,cf_rng_at,'=AND($H4="PENDENTE",ISNUMBER($F4),$F4<TODAY())',C_LRED,C_RED)
# P3: Prazo em 7 dias + PENDENTE → ORANGE
add_cf(ws,cf_rng_at,'=AND($H4="PENDENTE",ISNUMBER($F4),$F4>=TODAY(),$F4<=TODAY()+7)',C_LORANGE,C_ORANGE)
# P4: PENDENTE geral → YELLOW
add_cf(ws,cf_rng_at,'=$H4="PENDENTE"',C_LYELL,C_YELLOW)
# P5: EM ANDAMENTO → light blue
add_cf(ws,cf_rng_at,'=$H4="EM ANDAMENTO"',C_LBLUE,"1F3864")
# P6: CONCLUÍDO → soft GREEN
add_cf(ws,cf_rng_at,'=$H4="CONCLUÍDO"',"F0FFF0","375623")
# DataBar para % Conclusão — adicionado por último para coexistir com CFs acima
ws.conditional_formatting.add(f"G4:G{3+n_at}",
    DataBarRule(start_type='num',start_value=0,end_type='num',end_value=1,color="2E75B6"))

resp_str=",".join(RESP_UNICOS[:20]) if RESP_UNICOS else "Eventos,Time de Marca"
adv(ws,f'"{",".join(STATUS_AT)}"',f"H4:H{3+n_at}","Status")
adv(ws,f'"{resp_str}"',f"E4:E{3+n_at}","Responsável")

# Validação de DATA na coluna Prazo (bloqueia texto, permite vazio)
dv_prazo=DataValidation(type="date",operator="greaterThanOrEqual",formula1="43831",
    allow_blank=True,showInputMessage=True,showErrorMessage=True)
dv_prazo.promptTitle="Prazo da Atividade"
dv_prazo.prompt="Insira a data no formato DD/MM/AAAA.\nDeixe em branco se ainda não houver prazo definido."
dv_prazo.errorTitle="Data inválida"
dv_prazo.error="Insira uma data válida (DD/MM/AAAA) ou deixe em branco. Texto não é permitido."
dv_prazo.sqref=f"F4:F{3+n_at}"; ws.add_data_validation(dv_prazo)

print("DB_ATIVIDADES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# DB_FINANCEIRO + CF
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("DB_FINANCEIRO")
ws.sheet_view.showGridLines=False; ws.freeze_panes="C4"
tblock(ws,'A1:N1',"💰 BANCO DE DADOS FINANCEIRO  ·  v4  ·  Alertas de Pagamento","843C0C",sz=12)
ws.row_dimensions[1].height=28
sblock(ws,'A2:N2',
    "🟠 Linha laranja = Status Pag. Pendente  ·  🟡 Amarelo = A preencher  ·  "
    "🟢 Pago = verde  ·  PO=Pendente destaca em laranja",C_LORANGE,"843C0C")
ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=28

COLS_FIN=[
    (1,'ID','A',5),(2,'Evento','B',34),(3,'EM (Padrão)','C',14),(4,'FY','D',6),
    (5,'Fornecedor','E',22),(6,'Tipo de Custo','F',20),(7,'PO/WF','G',14),
    (8,'Orç. Previsto','H',16),(9,'Real. FY2025','I',14),(10,'Real. FY2026','J',14),
    (11,'Total Evento','K',16),(12,'Status Pag.','L',14),(13,'Status NF/ND','M',14),
    (14,'Observações','N',38),
]
for ci,label,col_l,w in COLS_FIN: ws.column_dimensions[get_column_letter(ci)].width=w
hdr(ws,3,range(1,15),bg="843C0C")
for ci,label,_,_ in COLS_FIN: ws.cell(row=3,column=ci).value=label

tab3=Table(displayName="TblFinanceiro",ref=f"A3:N{3+n_fin}")
tab3.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True)
ws.add_table(tab3)

total_prev_all=sum(f['Prev'] for f in financeiro)
total_r25=sum(f['R25'] for f in financeiro)
total_r26=sum(f['R26'] for f in financeiro)

for i,fin in enumerate(financeiro):
    r=4+i; drow(ws,r,14,alt=(i%2==1))
    row_data=[i+1,fin['Evento'],fin['EM'],fin['FY'],fin['Forn'],fin['Tipo'],fin['PO'],
              fin['Prev'],fin['R25'],fin['R26'],fin['Tot'],fin['StatPag'],fin['StatNF'],fin['Obs']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (8,9,10,11) and isinstance(val,(int,float)):
            cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
        if c_idx==12: scell(cell,fin['StatPag'])
        if c_idx==13: scell(cell,fin['StatNF'])

r_tot=4+n_fin; ws.row_dimensions[r_tot].height=22
ws.merge_cells(f'A{r_tot}:G{r_tot}')
c=ws.cell(row=r_tot,column=1); c.value="TOTAIS GERAIS"
c.fill=fx(C_NAVY); c.font=ft(True,"FFFFFF",10); c.alignment=al()
for col_idx,tot in [(8,total_prev_all),(9,total_r25),(10,total_r26),(11,total_r25+total_r26)]:
    cell=ws.cell(row=r_tot,column=col_idx); cell.value=tot
    cell.number_format='R$ #,##0.00'; cell.fill=fx(C_NAVY)
    cell.font=ft(True,"FFFFFF",10); cell.alignment=al("right"); cell.border=bd()

# ── CF rules — DB_FINANCEIRO ─────────────────────────────────────────────────
cf_rng_fin=f"A4:N{3+n_fin}"
# P1: Status Pag = Pendente → ORANGE row
add_cf(ws,cf_rng_fin,'=$L4="Pendente"',C_LORANGE,C_ORANGE)
# P2: Status Pag = A preencher → YELLOW
add_cf(ws,cf_rng_fin,'=$L4="A preencher"',C_LYELL,C_YELLOW)
# P3: PO = Pendente → YELLOW
add_cf(ws,cf_rng_fin,'=$G4="Pendente"',C_LYELL,C_YELLOW)
# P4: Pago → subtle green
add_cf(ws,cf_rng_fin,'=OR($L4="Pago",$L4="pago")',"F0FFF0","375623")

adv(ws,f'"{",".join(STATUS_PAG)}"',f"L4:L{3+n_fin}","Status Pagamento")
adv(ws,f'"{",".join(STATUS_NF)}"',f"M4:M{3+n_fin}","Status NF/ND/Recibo")
adv(ws,f'"{",".join(TIPO_CUSTO)}"',f"F4:F{3+n_fin}","Tipo de Custo")

print("DB_FINANCEIRO ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# DB_FORNECEDORES
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("DB_FORNECEDORES")
ws.sheet_view.showGridLines=False; ws.freeze_panes="B4"
tblock(ws,'A1:J1',"🏢 BANCO DE DADOS DE FORNECEDORES",C_TEAL,sz=13)
ws.row_dimensions[1].height=28
sblock(ws,'A2:J2',"Cadastro único · Atualizar CNPJ antes de criar PO · Homologados = verde · A verificar = amarelo","D9EAF5","1F6B75")
ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=32
COLS_FORN=[(1,'ID',5),(2,'Fornecedor',28),(3,'Categoria',22),(4,'Serviços',30),
    (5,'UF Sede',8),(6,'Status Homol.',16),(7,'CNPJ',18),(8,'Contato',22),
    (9,'E-mail',30),(10,'Observações',40)]
for ci,label,w in COLS_FORN: ws.column_dimensions[get_column_letter(ci)].width=w
hdr(ws,3,range(1,11),bg=C_TEAL)
for ci,label,_ in COLS_FORN: ws.cell(row=3,column=ci).value=label
forn_list=[(k,v[0],v[1],v[2],v[3]) for k,v in SUPPLIER_CATALOG.items()]
tab4=Table(displayName="TblFornecedores",ref=f"A3:J{3+len(forn_list)}")
tab4.tableStyleInfo=TableStyleInfo(name="TableStyleMedium9",showRowStripes=True)
ws.add_table(tab4)
for i,(name,cat,serv,uf,stat) in enumerate(forn_list):
    r=4+i; drow(ws,r,10,alt=(i%2==1))
    for ci,val in enumerate([i+1,name,cat,serv,uf,stat,'A preencher','A preencher','A preencher',''],1):
        cell=ws.cell(row=r,column=ci); cell.value=val
        if ci==6: scell(cell,val)
adv(ws,f'"{",".join(STATUS_FORN)}"',f"F4:F{3+len(forn_list)}","Status Homologação")
print("DB_FORNECEDORES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# DB_PARTICIPANTES
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("DB_PARTICIPANTES")
ws.sheet_view.showGridLines=False; ws.freeze_panes="B4"
tblock(ws,'A1:L1',"👥 BANCO DE DADOS DE PARTICIPANTES",C_PURPLE,sz=13)
ws.row_dimensions[1].height=28
sblock(ws,'A2:L2',"Registro histórico HCPs e Staff · Brain 2026 como base · Expandir para cada evento","EAD1FF","7030A0")
ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=32
COLS_P=[(1,'ID',5),(2,'Nome',30),(3,'Cargo/Regional',28),(4,'Tipo',16),
    (5,'Especialidade',22),(6,'CRM',14),(7,'UF',6),(8,'Evento Ref.',22),
    (9,'Hotel Bloco',16),(10,'Aéreo R$',12),(11,'Hospedagem R$',14),(12,'Observações',38)]
for ci,label,w in COLS_P: ws.column_dimensions[get_column_letter(ci)].width=w
hdr(ws,3,range(1,13),bg=C_PURPLE)
for ci,label,_ in COLS_P: ws.cell(row=3,column=ci).value=label
ws_ap=src['Brain Apoios - novo']
skip={'Hilton (Hotel 1)','Manhatam (Hotel 2)','Park Plaza (Hotel 3)','Participantes'}
parts=[]; seen=set()
for row in ws_ap.iter_rows(min_row=2,max_row=49,values_only=True):
    if row[0] is None: continue
    name=str(row[0]).strip()
    if name in skip or 'Apoios ' in name or not name or name in seen: continue
    reg=str(row[1]).strip() if row[1] else ''
    if reg in ('',None) or name.startswith('Apoio '): continue
    parts.append((name,f"Regional {reg}",'Staff Takeda','BRAIN 2026',f"Hotel {row[3]}" if row[3] else ''))
    seen.add(name)
tab5=Table(displayName="TblParticipantes",ref=f"A3:L{3+len(parts)}")
tab5.tableStyleInfo=TableStyleInfo(name="TableStyleMedium28",showRowStripes=True)
ws.add_table(tab5)
TIPO_PART='Staff Takeda,HCP Convidado,Palestrante,Coordenador,Moderador'
for i,(nome,cargo,tipo,evref,hotel) in enumerate(parts):
    r=4+i; drow(ws,r,12,alt=(i%2==1))
    for ci,val in enumerate([i+1,nome,cargo,tipo,'A preencher','A preencher','A preencher',evref,hotel,0,0,''],1):
        cell=ws.cell(row=r,column=ci); cell.value=val
        if ci in (10,11): cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
adv(ws,f'"{TIPO_PART}"',f"D4:D{3+len(parts)}","Tipo")
print("DB_PARTICIPANTES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# ⚡ ALERTAS  — fórmulas calculadas em tempo real
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("⚡ ALERTAS")
ws.sheet_view.showGridLines=False
for c,w in zip(range(1,10),[2,28,20,20,20,20,20,20,2]):
    ws.column_dimensions[get_column_letter(c)].width=w

ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=36
ws.merge_cells('B2:H2')
c=ws['B2']; c.value="⚡ PAINEL DE ALERTAS EM TEMPO REAL"
c.fill=fx("C4700E"); c.font=ft(True,"FFFFFF",18); c.alignment=al()
ws.row_dimensions[3].height=18; ws.merge_cells('B3:H3')
c=ws['B3']; c.value=f"Atualizado: {datetime.date.today().strftime('%d/%m/%Y')}  ·  Formulas referenciam TblEventos / TblAtividades / TblFinanceiro  ·  Requer Excel 365 para FILTER"
c.fill=fx("FCE4D6"); c.font=ft(False,"843C0C",9); c.alignment=al("left")

def alert_section(ws, r, col_s, col_e, title, bg, items):
    """Write an alert section with label/formula pairs."""
    ws.row_dimensions[r].height=22
    ws.merge_cells(f"{get_column_letter(col_s)}{r}:{get_column_letter(col_e)}{r}")
    c=ws.cell(row=r,column=col_s); c.value=title
    c.fill=fx(bg); c.font=ft(True,"FFFFFF",11); c.alignment=al()
    r+=1
    for label,formula,fmt in items:
        ws.row_dimensions[r].height=20
        c_label=ws.cell(row=r,column=col_s)
        c_label.value=label; c_label.fill=fx(C_LGRAY)
        c_label.font=ft(False,"404040",9); c_label.alignment=al("left"); c_label.border=bd()
        ws.merge_cells(f"{get_column_letter(col_s+1)}{r}:{get_column_letter(col_s+2)}{r}")
        c_val=ws.cell(row=r,column=col_s+1)
        c_val.value=formula; c_val.fill=fx(C_WHITE)
        c_val.font=ft(True,"1F3864",12); c_val.alignment=al(); c_val.border=bd()
        if fmt=='money': c_val.number_format='R$ #,##0'
        r+=1
    return r

r=5
# ── CRÍTICO (red) ─────────────────────────────────────────────────────────────
r=alert_section(ws,r,2,8,"🔴 CRÍTICO — Requer ação imediata","C00000",[
    ("Eventos VENCIDOS (a Realizar + data passada)",
     '=COUNTIFS(TblEventos[Status Evento],"a Realizar",TblEventos[Data Início],"<"&TODAY(),TblEventos[Data Início],"<>")',None),
    ("Atividades c/ PRAZO VENCIDO (PENDENTE + data<hoje)",
     '=COUNTIFS(TblAtividades[Status],"PENDENTE",TblAtividades[Prazo],"<"&TODAY())',None),
    ("Pagamentos PENDENTES (sem PO definido)",
     '=COUNTIF(TblFinanceiro[PO/WF],"Pendente")',None),
    ("EMs DUPLICADOS detectados",
     '=COUNTIF(TblEventos[Flag],"*dup*")',None),
])
r+=1
# ── ATENÇÃO (orange) ──────────────────────────────────────────────────────────
r=alert_section(ws,r,2,8,"🟠 ATENÇÃO — Monitorar esta semana","C4700E",[
    ("Eventos nos próximos 7 dias",
     '=COUNTIFS(TblEventos[Status Evento],"a Realizar",TblEventos[Data Início],">="&TODAY(),TblEventos[Data Início],"<="&(TODAY()+7))',None),
    ("Atividades com prazo em 7 dias",
     '=COUNTIFS(TblAtividades[Status],"PENDENTE",TblAtividades[Prazo],">="&TODAY(),TblAtividades[Prazo],"<="&(TODAY()+7))',None),
    ("Atividades PENDENTES total",
     '=COUNTIF(TblAtividades[Status],"PENDENTE")',None),
    ("NFs/Recibos a preencher",
     '=COUNTIF(TblFinanceiro[Status NF/ND],"A preencher")',None),
])
r+=1
# ── MONITORAR (yellow) ────────────────────────────────────────────────────────
r=alert_section(ws,r,2,8,"🟡 MONITORAR — Próximos 30 dias","7F6000",[
    ("Eventos nos próximos 30 dias",
     '=COUNTIFS(TblEventos[Status Evento],"a Realizar",TblEventos[Data Início],">="&TODAY(),TblEventos[Data Início],"<="&(TODAY()+30))',None),
    ("Atividades EM ANDAMENTO",
     '=COUNTIF(TblAtividades[Status],"EM ANDAMENTO")',None),
    ("Eventos sem responsável (Solicitante vazio)",
     '=COUNTIF(TblEventos[Solicitante],"")',None),
])
r+=1
# ── FINANCEIRO ────────────────────────────────────────────────────────────────
r=alert_section(ws,r,2,8,"💰 RESUMO FINANCEIRO","375623",[
    ("Budget total previsto",
     '=SUM(TblFinanceiro[Orç. Previsto])',  'money'),
    ("Total realizado FY2025",
     '=SUM(TblFinanceiro[Real. FY2025])',   'money'),
    ("Total realizado FY2026",
     '=SUM(TblFinanceiro[Real. FY2026])',   'money'),
    ("Pagamentos Pendentes (budget estimado)",
     '=SUMIF(TblFinanceiro[Status Pag.],"Pendente",TblFinanceiro[Orç. Previsto])', 'money'),
    ("% Budget executado (FY25+FY26)",
     '=IFERROR((SUM(TblFinanceiro[Real. FY2025])+SUM(TblFinanceiro[Real. FY2026]))/SUM(TblFinanceiro[Orç. Previsto]),0)',None),
])
# Format % row
ws.cell(row=r-1,column=3).number_format='0.0%'
r+=1

# ── Dynamic lists (FILTER — requires Excel 365) ───────────────────────────────
ws.row_dimensions[r].height=22; ws.merge_cells(f'B{r}:H{r}')
c=ws.cell(row=r,column=2)
c.value="📋 LISTAS DINÂMICAS — Requerem Excel 365 (função FILTER)"
c.fill=fx(C_NAVY); c.font=ft(True,"FFFFFF",11); c.alignment=al(); r+=1

ws.row_dimensions[r].height=18; ws.merge_cells(f'B{r}:H{r}')
c=ws.cell(row=r,column=2)
c.value="Se a fórmula mostrar #NOME? ou #VALOR!, seu Excel não suporta FILTER. Use os filtros nativos das tabelas DB_."
c.fill=fx(C_LYELL); c.font=ft(False,C_YELLOW,9); c.alignment=al("left"); r+=2

filter_sections=[
    ("🔴 EVENTOS VENCIDOS (a Realizar com data passada)","C00000",
     '=IFERROR(FILTER(TblEventos[Evento],(TblEventos[Status Evento]="a Realizar")*(TblEventos[Data Início]<TODAY())*(ISNUMBER(TblEventos[Data Início]))),"✅ Nenhum evento vencido")'),
    ("🟠 EVENTOS PRÓXIMOS 30 DIAS","C4700E",
     '=IFERROR(FILTER(TblEventos[Evento],(TblEventos[Status Evento]="a Realizar")*(ISNUMBER(TblEventos[Data Início]))*(TblEventos[Data Início]>=TODAY())*(TblEventos[Data Início]<=TODAY()+30)),"✅ Nenhum evento nos próximos 30 dias")'),
    ("💰 PAGAMENTOS PENDENTES","843C0C",
     '=IFERROR(FILTER(TblFinanceiro[Evento],TblFinanceiro[Status Pag.]="Pendente"),"✅ Nenhum pagamento pendente")'),
    ("✅ ATIVIDADES PENDENTES (amostra)","375623",
     '=IFERROR(FILTER(TblAtividades[Atividade],TblAtividades[Status]="PENDENTE"),"✅ Nenhuma atividade pendente")'),
]
for title,bg,formula in filter_sections:
    ws.row_dimensions[r].height=20; ws.merge_cells(f'B{r}:H{r}')
    c=ws.cell(row=r,column=2); c.value=title
    c.fill=fx(bg); c.font=ft(True,"FFFFFF",10); c.alignment=al(); r+=1
    # Formula cell (spill range — leave 20 rows)
    cell=ws.cell(row=r,column=2); cell.value=formula
    cell.font=ft(False,"1F3864",9); cell.alignment=al("left"); cell.border=bd()
    ws.merge_cells(f'B{r}:H{r}')
    r+=22  # space for spill

print("⚡ ALERTAS ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("📊 DASHBOARD")
ws.sheet_view.showGridLines=False
for c in range(1,16): ws.column_dimensions[get_column_letter(c)].width=14
ws.column_dimensions['A'].width=2
ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=40
ws.merge_cells('B2:O2'); c=ws['B2']; c.value="📊 DASHBOARD EXECUTIVO · v4 · Fase 3"
c.fill=fx(C_NAVY); c.font=ft(True,"FFFFFF",20); c.alignment=al()
ws.merge_cells('B3:O3'); c=ws['B3']
c.value=f"Atualizado: {datetime.date.today().strftime('%d/%m/%Y')}  ·  Alertas visuais ativos nas abas DB_  ·  Ver ⚡ ALERTAS para lista de pendências"
c.fill=fx(C_LBLUE); c.font=ft(False,C_NAVY,10); c.alignment=al("left")
ws.row_dimensions[3].height=18
ws.row_dimensions[5].height=18; ws.row_dimensions[6].height=50; ws.row_dimensions[7].height=22
kpis=[
    ("TOTAL\nEVENTOS",n_ev,"",C_NAVY),
    ("EVENTOS\n2026",sum(1 for e in events if e['Ano']=='2026'),"",C_BLUE),
    ("A\nREALIZAR",sum(1 for e in events if 'Realizar' in e['StatusEv']),"","843C0C"),
    ("REALIZADOS",sum(1 for e in events if e['StatusEv']=='Realizado'),"","375623"),
    ("PENDENTES",sum(1 for a in activities if a['Status']=='PENDENTE'),"",C_RED),
    (f"BUDGET",f"R$ {total_prev_all/1e6:.1f}M","",C_PURPLE),
]
kpi_cols=[2,4,6,8,10,12]
for idx,(label,value,sub,bg) in enumerate(kpis):
    col=kpi_cols[idx]; col2=col+1
    for row_r in [5,6,7]:
        ws.merge_cells(start_row=row_r,start_column=col,end_row=row_r,end_column=col2)
    for rr,val,sz in [(5,label,9),(6,value,26),(7,sub,8)]:
        c=ws.cell(row=rr,column=col); c.value=val
        c.fill=fx(bg) if rr!=7 else fx(C_LGRAY)
        c.font=ft(True,"FFFFFF" if rr!=7 else "404040",sz)
        c.alignment=al(); c.border=bd()

status_count=Counter(e['StatusEv'] for e in events)
bu_count=Counter(e['BU'] for e in events if e['BU'])
city_count=Counter(e['Cidade'] for e in events if e['Cidade'])
act_status=Counter(a['Status'] for a in activities)
pay_status=Counter(f['StatPag'] for f in financeiro if f['StatPag'])
fin_by_ev=defaultdict(lambda:{'prev':0,'r25':0,'r26':0})
for f in financeiro:
    fin_by_ev[f['Evento']]['prev']+=f['Prev']
    fin_by_ev[f['Evento']]['r25']+=f['R25']
    fin_by_ev[f['Evento']]['r26']+=f['R26']

def tbl(ws,rs,cs,title,tbg,cols_h,rows):
    ce=cs+len(cols_h)-1
    ws.merge_cells(start_row=rs,start_column=cs,end_row=rs,end_column=ce)
    c=ws.cell(row=rs,column=cs); c.value=title; c.fill=fx(tbg)
    c.font=ft(True,"FFFFFF",11); c.alignment=al()
    for ci,h in enumerate(cols_h,cs):
        c=ws.cell(row=rs+1,column=ci); c.value=h
        c.fill=fx(tbg); c.font=ft(True,"FFFFFF",9); c.alignment=al(); c.border=bd()
    for ri,row_vals in enumerate(rows):
        r=rs+2+ri
        for ci,val in enumerate(row_vals,cs):
            c=ws.cell(row=r,column=ci); c.value=val
            c.fill=fx(C_LGRAY if r%2==0 else C_WHITE); c.font=ft(size=9)
            c.alignment=al(); c.border=bd()
    return rs+2+len(rows)

r_s=9
tbl(ws,r_s,2,"STATUS DOS EVENTOS",C_NAVY,['Status','Qtd','%'],
    [(s,cnt,f"{cnt/n_ev*100:.0f}%") for s,cnt in sorted(status_count.items(),key=lambda x:-x[1])])
tbl(ws,r_s,6,"POR BU","375623",['BU','Qtd'],
    list(sorted(bu_count.items(),key=lambda x:-x[1]))[:8])
tbl(ws,r_s,9,"TOP CIDADES",C_TEAL,['Cidade','Qtd'],
    list(sorted(city_count.items(),key=lambda x:-x[1]))[:8])
tbl(ws,r_s,12,"PAGAMENTOS",C_PURPLE,['Status','Qtd'],
    list(sorted(pay_status.items(),key=lambda x:-x[1])))
r_s2=22
tbl(ws,r_s2,2,"ATIVIDADES POR STATUS",C_NAVY,['Status','Qtd','%'],
    [(s,cnt,f"{cnt/n_at*100:.0f}%") for s,cnt in sorted(act_status.items(),key=lambda x:-x[1])])
end_r=tbl(ws,r_s2,6,"RESUMO FINANCEIRO","843C0C",['Evento','Previsto','Realizado'],
    [(k[:28],v['prev'],v['r25']+v['r26']) for k,v in sorted(fin_by_ev.items(),key=lambda x:-x[1]['prev'])[:10]])
for r in range(r_s2+2,end_r):
    for col in [7,8]:
        cell=ws.cell(row=r,column=col)
        if isinstance(cell.value,(int,float)): cell.number_format='R$ #,##0'; cell.alignment=al("right")
print("DASHBOARD ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# PARTICIPAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("PARTICIPAÇÃO")
ws.sheet_view.showGridLines=False; ws.freeze_panes="B4"
tblock(ws,'A1:M1',"✈️ CONTROLE DE PARTICIPAÇÃO E LOGÍSTICA",C_NAVY,sz=13)
ws.row_dimensions[1].height=28
sblock(ws,'A2:M2',"Custos de deslocamento e hospedagem por participante",C_LBLUE,C_NAVY)
ws.row_dimensions[2].height=16
COLS_PT=[(1,'Evento',34),(2,'Participante',22),(3,'Cargo',18),(4,'Regional',10),
    (5,'Cidade Orig.',14),(6,'Data Ida',12),(7,'Data Volta',12),
    (8,'Aéreo R$',12),(9,'Hospedagem R$',14),(10,'Refeição R$',12),
    (11,'Deslocamento R$',16),(12,'Total R$',14),(13,'Observações',30)]
for ci,label,w in COLS_PT: ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[3].height=28; hdr(ws,3,range(1,14))
for ci,label,_ in COLS_PT: ws.cell(row=3,column=ci).value=label
ws_p_src=src['PARTICIPAÇÃO']; r_p=4
for row in ws_p_src.iter_rows(min_row=3,max_row=ws_p_src.max_row,values_only=True):
    if row[1] is None: continue
    drow(ws,r_p,13,alt=(r_p%2==0))
    def to_dt(v): return v.date() if isinstance(v,datetime.datetime) else v
    def to_n(v):
        try: return float(str(v)) if v else 0
        except: return 0
    aer=to_n(row[6]); hosp=to_n(row[7]); ref=to_n(row[8]); desl=to_n(row[9])
    vals=[str(row[1]).strip(),'','','','',to_dt(row[4]),to_dt(row[5]),aer,hosp,ref,desl,aer+hosp+ref+desl,'']
    for ci,val in enumerate(vals,1):
        cell=ws.cell(row=r_p,column=ci); cell.value=val
        if ci in (8,9,10,11,12): cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
        if ci in (6,7) and isinstance(val,datetime.date): cell.number_format='DD/MM/YYYY'
    r_p+=1
tab6=Table(displayName="TblParticipacao",ref=f"A3:M{r_p-1}")
tab6.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tab6)
print("PARTICIPAÇÃO ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# MODELO_CHECKLIST (com coluna Prazo)
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("MODELO_CHECKLIST")
ws.sheet_view.showGridLines=False
for c,w in zip(range(1,9),[2,25,38,18,13,13,16,40]):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=36
ws.merge_cells('B2:H2'); c=ws['B2']
c.value="📝 MODELO PADRÃO DE CHECKLIST  ·  v4 · com coluna Prazo"
c.fill=fx(C_NAVY); c.font=ft(True,"FFFFFF",13); c.alignment=al()
ws.row_dimensions[3].height=20; ws.merge_cells('B3:H3'); c=ws['B3']
c.value=("✅ NOVO: Coluna Prazo habilita alertas automáticos de CF  ·  "
         "Duplicar aba → renomear → preencher evento → copiar dados para DB_ATIVIDADES")
c.fill=fx(C_LYELL); c.font=ft(False,C_YELLOW,9); c.alignment=al("left")
ws.row_dimensions[5].height=22
for col,(label) in enumerate(['Evento:','EM:','Data:'],2):
    c=ws.cell(row=5,column=col); c.value=label
    c.fill=fx(C_LBLUE); c.font=ft(True,C_NAVY,9); c.alignment=al(); c.border=bd()
    c2=ws.cell(row=5,column=col+1); c2.value='← preencher'
    c2.fill=fx(C_LYELL); c2.font=ft(False,C_YELLOW,9); c2.alignment=al("left"); c2.border=bd()
ws.row_dimensions[7].height=28
hdr(ws,7,[2,3,4,5,6,7,8])
for ci,h in enumerate(['Etapa','Atividade','Responsável','Prazo','% Conclusão','Status','Observações'],2):
    ws.cell(row=7,column=ci).value=h
TMPL=[
    ('Planejamento','Briefing com Time de Marca','Time de Marca',None,0,'PENDENTE',''),
    ('Planejamento','Inclusão no calendário de eventos','Eventos',None,0,'PENDENTE',''),
    ('VEEVA','EM aprovada (Cota)','Time de Marca',None,0,'PENDENTE','Número EM: ___'),
    ('VEEVA','EM aprovada (Apoio)','Time de Marca',None,0,'PENDENTE',''),
    ('VEEVA','EM aprovada (Simpósio)','Time de Marca',None,0,'PENDENTE',''),
    ('VEEVA','Incluir time de Processos/Pagamentos','Eventos',None,0,'PENDENTE',''),
    ('VEEVA','Incluir Agência Logística na EM','Eventos',None,0,'PENDENTE',''),
    ('Sociedade','Contrato de Cota recebido','Eventos',None,0,'PENDENTE',''),
    ('Sociedade','Envio de logo para Sociedade','Eventos',None,0,'PENDENTE',''),
    ('Sociedade','Monitorar contrapartidas','Eventos',None,0,'PENDENTE',''),
    ('Agência','Briefing para agência logística','Eventos',None,0,'PENDENTE',''),
    ('Agência','Solicitar orçamento','Eventos',None,0,'PENDENTE',''),
    ('Agência','Aprovação do orçamento','Time de Marca',None,0,'PENDENTE',''),
    ('Agência','Criação de PO','Eventos',None,0,'PENDENTE','PO Nº: ___'),
    ('Agência','RSVP / Inscrições HCPs','Eventos',None,0,'PENDENTE',''),
    ('Simpósio','Contratos de Palestrantes','Eventos',None,0,'PENDENTE',''),
    ('Simpósio','PO Palestrantes aprovada','Eventos',None,0,'PENDENTE',''),
    ('Simpósio','Aditivo (data/horário corretos)','Eventos',None,0,'PENDENTE',''),
    ('Simpósio','Lunch Box / Buffet','Eventos',None,0,'PENDENTE',''),
    ('Stand','Documentação do expositor','Eventos',None,0,'PENDENTE',''),
    ('Stand','Formulários obrigatórios','Eventos',None,0,'PENDENTE',''),
    ('Stand','PO Montadora aprovada','Eventos',None,0,'PENDENTE',''),
    ('Pós-Evento','Solicitação NF/Recibo','Eventos',None,0,'PENDENTE',''),
    ('Pós-Evento','Confirmação de pagamento','Eventos',None,0,'PENDENTE',''),
    ('Pós-Evento','Encerramento da EM','Eventos',None,0,'PENDENTE',''),
    ('Pós-Evento','Relatório de execução','Eventos',None,0,'PENDENTE',''),
    ('Pós-Evento','Atualizar DB_ATIVIDADES','Eventos',None,0,'PENDENTE','Copiar linhas concluídas'),
]
for i,(etapa,ativ,resp,prazo,pct,stat,obs) in enumerate(TMPL):
    r=8+i; ws.row_dimensions[r].height=18; alt=(i%2==1)
    for col in range(2,9): ws.cell(row=r,column=col).fill=fx(C_LGRAY if alt else C_WHITE)
    vals=[etapa,ativ,resp,'A definir',pct/100 if pct else 0,stat,obs]
    for c_idx,val in enumerate(vals,2):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        cell.font=ft(size=9); cell.border=bd()
        if c_idx==6: scell(cell,val)
        elif c_idx==5: cell.number_format='0%'; cell.alignment=al()
        elif c_idx==4: cell.font=ft(False,"7F6000",8,italic=True); cell.alignment=al("center")
        else: cell.alignment=al("left")
        if c_idx==2:
            is_new_etapa=(i==0 or etapa!=TMPL[i-1][0])
            if is_new_etapa:
                cell.fill=fx(C_LBLUE); cell.font=ft(True,C_NAVY,9)
adv(ws,f'"{",".join(STATUS_AT)}"',f"G8:G{7+len(TMPL)}","Status")
print("MODELO_CHECKLIST ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# ROADMAP (Fase 3 concluída)
# ═══════════════════════════════════════════════════════════════════════════════
ws=wb.create_sheet("🗺️ ROADMAP")
ws.sheet_view.showGridLines=False
for c in range(1,10): ws.column_dimensions[get_column_letter(c)].width=22
ws.column_dimensions['A'].width=3
ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=40
ws.merge_cells('B2:I2'); c=ws['B2']; c.value="🗺️ ROADMAP DE IMPLEMENTAÇÃO"
c.fill=fx(C_NAVY); c.font=ft(True,"FFFFFF",20); c.alignment=al()
phases=[
    ("FASE 1 — QUICK WINS ✅","375623","E2EFDA","375623",[
        ("✅","DB_EVENTOS, DB_ATIVIDADES, DB_FINANCEIRO centralizados"),
        ("✅","Excel Tables com filtros e dropdown validations"),
        ("✅","Dashboard executivo com KPIs"),
        ("✅","Formatação condicional de status"),
    ]),
    ("FASE 2 — REESTRUTURAÇÃO ✅",C_BLUE,"BDD7EE","1F3864",[
        ("✅","DB_FORNECEDORES: 17 fornecedores cadastrados"),
        ("✅","DB_PARTICIPANTES: registro histórico de HCPs"),
        ("✅","15 validações de dados (dropdowns)"),
        ("✅","ID_Unico padronizado + flag de EMs duplicados"),
        ("✅","MODELO_CHECKLIST: 27 atividades padrão"),
        ("✅","Correções: typos, EMs normalizados, campos vazios"),
    ]),
    ("FASE 3 — AUTOMAÇÃO ✅","C4700E","FCE4D6","843C0C",[
        ("✅","CF Visual Alerts: DB_EVENTOS (vencidos/7d/30d) + DB_ATIVIDADES (prazo+PENDENTE) + DB_FINANCEIRO (pendente/pago)"),
        ("✅","Coluna Prazo em DB_ATIVIDADES — habilita alertas por deadline"),
        ("✅","Aba ⚡ ALERTAS com COUNTIFS+SUMIFS e FILTER (Excel 365)"),
        ("✅","Office Script: automation/office_script_nec.ts — relatório semanal gerado automaticamente"),
        ("✅","Power Query: automation/power_query_nec.pq — 4 queries M prontos para conexão"),
        ("✅","Power Automate: automation/pa_flows_nec.md — 4 fluxos documentados e prontos para criação"),
        ("✅","SharePoint: automation/sharepoint_arquitetura.md — arquitetura e governança completa"),
        ("⚠️ BLOQUEIO","Rede egress bloqueia graph.microsoft.com — PA/SPO não criados via API automaticamente"),
        ("⚠️ BLOQUEIO","OAuth2/Service Principal não configurado no ambiente de execução"),
        ("📋 AÇÃO","IT: adicionar graph.microsoft.com ao allowlist de egress → automação pode ser finalizada"),
    ]),
    ("FASE 4 — DASHBOARD AVANÇADO",C_PURPLE,"EAD1FF","7030A0",[
        ("📅","Power BI Desktop conectado às TblEventos/TblAtividades/TblFinanceiro"),
        ("📅","Gantt de eventos por trimestre (visual timeline)"),
        ("📅","Mapa geográfico: eventos por cidade brasileira"),
        ("📅","Análise YoY: budget x realizado 2025 vs 2026"),
        ("📅","Relatório executivo PDF automático (Power BI + PA)"),
        ("📅","Scorecard de performance por responsável/BU"),
    ]),
]
r=4
for phase_title,bg_hdr,bg_light,fg_dark,items in phases:
    ws.row_dimensions[r].height=28
    ws.merge_cells(f'B{r}:I{r}')
    c=ws.cell(row=r,column=2); c.value=phase_title
    c.fill=fx(bg_hdr); c.font=ft(True,"FFFFFF",12); c.alignment=al(); r+=1
    for stat,desc in items:
        ws.row_dimensions[r].height=20
        c=ws.cell(row=r,column=2); c.value=stat
        bg_s=("E2EFDA" if stat=="✅" else "FFDDD8" if "BLOQUEIO" in stat else
              "FFF2CC" if "AÇÃO" in stat else "F2F2F2")
        c.fill=fx(bg_s); c.font=ft(True,fg_dark,9); c.alignment=al(); c.border=bd()
        ws.merge_cells(f'C{r}:I{r}')
        c=ws.cell(row=r,column=3); c.value=desc
        c.fill=fx(C_WHITE if r%2==0 else C_LGRAY); c.font=ft(size=9)
        c.alignment=al("left"); c.border=bd(); r+=1
    r+=1

print("ROADMAP ✅")

# ── Tab colors ─────────────────────────────────────────────────────────────────
TAB={
    "🏠 INÍCIO":"1F3864","DB_EVENTOS":"2E75B6","DB_ATIVIDADES":"375623",
    "DB_FINANCEIRO":"843C0C","DB_FORNECEDORES":"1F6B75","DB_PARTICIPANTES":"7030A0",
    "⚡ ALERTAS":"C4700E","📊 DASHBOARD":"1F3864",
    "PARTICIPAÇÃO":"2E75B6","MODELO_CHECKLIST":"C4700E","🗺️ ROADMAP":"404040",
}
for sname,color in TAB.items():
    if sname in wb.sheetnames: wb[sname].sheet_properties.tabColor=color

out='/home/user/Femartignon/NEC_Eventos_Reestruturado_v4.xlsx'
wb.save(out)
print(f"\n✅ Excel v4 salvo → {out}")
