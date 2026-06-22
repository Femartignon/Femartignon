import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabel
import datetime

# ── Load source ───────────────────────────────────────────────────────────────
src = load_workbook('/root/.claude/uploads/6d41cea8-3482-52e3-859c-cdd7002a8a66/1318286f-Calenda_rio_Neuro_Acesso_NEC.xlsx', data_only=True)

# ── Styles ────────────────────────────────────────────────────────────────────
def fx(hex_color): return PatternFill("solid", fgColor=hex_color)
def ft(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def bd():
    s = Side(style="thin", color="D9D9D9"); return Border(left=s,right=s,top=s,bottom=s)
def bdM():
    s = Side(style="medium", color="2E75B6"); return Border(left=s,right=s,top=s,bottom=s)
def al(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

C_NAVY,C_BLUE,C_LBLUE  = "1F3864","2E75B6","BDD7EE"
C_LGRAY,C_WHITE         = "F2F2F2","FFFFFF"
C_GREEN,C_LGREEN        = "375623","E2EFDA"
C_ORANGE,C_LORANGE      = "843C0C","FCE4D6"
C_RED,C_LRED            = "C00000","FFDDD8"
C_YELLOW,C_LYELL        = "7F6000","FFF2CC"
C_PURPLE                = "7030A0"

TIPO_MAP = {'C':'Cota','A':'Apoio','S':'Simpósio','T':'TOME/Advisory'}

STATUS_FX = {
    'Realizado':    ('E2EFDA','375623'),
    'a Realizar':   ('BDD7EE','1F3864'),
    'Cancelado':    ('FFDDD8','C00000'),
    'CONCLUÍDO':    ('E2EFDA','375623'),
    'EM ANDAMENTO': ('FFF2CC','7F6000'),
    'PENDENTE':     ('FCE4D6','843C0C'),
    'ATRASADO':     ('FFDDD8','C00000'),
    '-':            ('F2F2F2','404040'),
    'Pago':         ('E2EFDA','375623'),
    'Pós Evento':   ('FFF2CC','7F6000'),
    'Pendente':     ('FCE4D6','843C0C'),
    'aprovado':     ('BDD7EE','1F3864'),
    'pago':         ('E2EFDA','375623'),
}

def hdr(ws, r, cols, bg=C_NAVY, fg=C_WHITE, sz=10):
    for c in cols:
        cell = ws.cell(row=r, column=c)
        cell.fill = fx(bg); cell.font = ft(True,fg,sz)
        cell.alignment = al(); cell.border = bd()

def data_row(ws, r, max_c, alt=False):
    bg = C_LGRAY if alt else C_WHITE
    for c in range(1, max_c+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fx(bg); cell.font = ft(size=9)
        cell.alignment = al("left"); cell.border = bd()

def status_cell(cell, status):
    key = status.strip() if status else ''
    if key in STATUS_FX:
        bg, fg = STATUS_FX[key]
        cell.fill = fx(bg); cell.font = ft(True, fg, 9)
    cell.alignment = al(); cell.border = bd()

def set_col_width(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def freeze(ws, cell): ws.freeze_panes = cell

# ─────────────────────────────────────────────────────────────────────────────
# Extract data
# ─────────────────────────────────────────────────────────────────────────────
cal = src['Calendário de Eventos']
events = []
for row in cal.iter_rows(min_row=4, max_row=60, values_only=True):
    if row[2] is None: continue
    name = str(row[2]).strip()
    if name in ('APOIO','COTA','SIMPÓSIO','TOME','','-'): continue
    def s(v): return str(v).strip() if v is not None else ''
    def d(v):
        if isinstance(v, datetime.datetime): return v.date()
        if isinstance(v, datetime.date): return v
        return None
    try: psp = int(float(str(row[19]))) if row[19] not in (None,'','-','X') else None
    except: psp = None
    try: psa = int(float(str(row[20]))) if row[20] not in (None,'','-','X') else None
    except: psa = None
    events.append({
        'Evento':s(row[2]),'Tipo':s(row[1]),'EM':s(row[4]),
        'Status Evento':s(row[7]),'Status EM':s(row[8]),
        'Solicitante':s(row[9]),'BU':s(row[10]),'Produto':s(row[11]),
        'Ano':s(row[12]),'Data Início':d(row[13]),'Data Fim':d(row[14]),
        'Cidade':s(row[16]),'Formato':s(row[17]),'Local':s(row[18]),
        'PSs Planejados':psp,'PSs Presentes':psa,
        'Agência':s(row[21]),'Hotel':s(row[22]),'Sociedade':s(row[23]),
        'Stand':s(row[24]),'Complexidade':s(row[3]),'Observações':s(row[27])
    })

# Activities
ACT_CFG = {
    'Fórum de Acesso 2026':  ('FÓRUM ACESSO 2026', 1,2,3,4,5,6, 5),
    ' FÓRUM GEDIIB ':        ('FORUM GEDIIB BRASÍLIA', 3,4,5,6,7,8, 5),
    'Cong. Medicina do Sono':('Congresso Medicina do Sono', 1,2,3,4,5,6, 5),
    'Brain 2026':            ('BRAIN 2026', 2,3,4,5,6,7, 6),
    'Brain Simpósio':        ('BRAIN 2026 - Simpósio', 2,3,4,5,6,7, 5),
    'Board Albert Einstein': ('Board Review Albert Einstein', 1,2,3,4,5,6, 5),
    'CONASENS':              ('CONASENS', 1,2,3,4,5,6, 5),
    'Aprender Criança 2026': ('APRENDER CRIANÇA 2026', 1,2,3,4,5,6, 5),
    'ADV. BOARD FRUZAQLA':   ('Advisory Board Fruzaqla', 1,2,3,4,5,6, 5),
    'Masterclass 2026':      ('MASTERCLASS 2026', 2,3,4,5,6,7, 5),
    'Masterclass Simpósio':  ('MASTERCLASS 2026 - Simpósio', 2,3,4,5,6,7, 5),
    'ABENEPI':               ('ABENEPI 2026', 1,2,3,4,5,6, 6),
}
activities = []
for sname, (ev, ce, ca, cr, cp, cs, co, start) in ACT_CFG.items():
    ws2 = src[sname]; last_e = ''
    for row in ws2.iter_rows(min_row=start, max_row=ws2.max_row, values_only=True):
        raw = list(row) + [None]*15
        ev_v=raw[ce-1]; av=raw[ca-1]; rv=raw[cr-1]; pv=raw[cp-1]; sv=raw[cs-1]; ov=raw[co-1]
        if ev_v not in (None,'') and str(ev_v).strip() not in ('Atividade','Responsável','Andamento (%)','Status','Observações'):
            last_e = str(ev_v).strip()
        if av is None or str(av).strip() in ('','Atividade','Responsável','Andamento (%)','Status','Observações'): continue
        try: pct = int(float(str(pv))) if pv is not None else 0
        except: pct = 0
        stat = str(sv).strip() if sv else ('-' if pct == 0 else 'EM ANDAMENTO')
        if not stat: stat = 'CONCLUÍDO' if pct == 100 else ('EM ANDAMENTO' if pct > 0 else 'PENDENTE')
        activities.append({'Evento':ev,'Etapa':last_e,'Atividade':str(av).strip(),
            'Responsável':str(rv).strip() if rv else '','% Conclusão':pct,
            'Status':stat,'Observações':str(ov).strip() if ov else ''})

# Financial
fin_ws = src['FINANCEIRO']
financeiro = []
for row in fin_ws.iter_rows(min_row=2, max_row=fin_ws.max_row, values_only=True):
    if row[0] is None: continue
    def n(v):
        if v is None: return 0.0
        try: return float(str(v).replace(',','.'))
        except: return 0.0
    financeiro.append({'Evento':str(row[0]).strip(),'EM':str(row[1]).strip() if row[1] else '',
        'FY':str(row[2]).strip() if row[2] else '','Fornecedor':str(row[3]).strip() if row[3] else '',
        'Tipo de Custo':str(row[4]).strip() if row[4] else '',
        'PO/WF':str(row[5]).strip() if row[5] else '',
        'Orçamento Previsto':n(row[6]),'Realizado FY2025':n(row[7]),'Realizado FY2026':n(row[8]),
        'Total Evento':n(row[9]),'Status Pagamento':str(row[10]).strip() if row[10] else '',
        'Status NF/ND':str(row[11]).strip() if row[11] else '',
        'Observações':str(row[12]).strip() if row[12] else ''})

print(f"Eventos: {len(events)} | Atividades: {len(activities)} | Financeiro: {len(financeiro)}")

# ═════════════════════════════════════════════════════════════════════════════
# Build new workbook
# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ─── Sheet order ──────────────────────────────────────────────────────────────
# 1. INICIO  2. DB_EVENTOS  3. DB_ATIVIDADES  4. DB_FINANCEIRO  5. DASHBOARD
# 6. PARTICIPAÇÃO  7. ROADMAP

# ═════════════════════════════════════════════════════════════════════════════
# 1. INICIO - Capa
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🏠 INÍCIO")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 20
ws.row_dimensions[1].height = 15

# Title block
for r in range(2, 8):
    ws.row_dimensions[r].height = 22
for c in range(2, 9):
    for r in range(2, 8):
        ws.cell(row=r, column=c).fill = fx(C_NAVY)

ws.merge_cells('B2:H4')
c = ws['B2']
c.value = "GESTÃO DE EVENTOS FARMACÊUTICOS"
c.font = ft(True, C_WHITE, 22); c.alignment = al()

ws.merge_cells('B5:H5')
c = ws['B5']
c.value = "Neurociências & Acesso  |  NEC  |  Takeda"
c.font = ft(False, "BDD7EE", 13); c.alignment = al()

ws.merge_cells('B6:H7')
c = ws['B6']
c.value = f"Versão Reestruturada  ·  {datetime.date.today().strftime('%d/%m/%Y')}"
c.font = ft(False, "9DC3E6", 10); c.alignment = al()

# Navigation cards
nav = [
    ("B", "DB_EVENTOS", "📋 Banco de Dados\nde Eventos", C_BLUE, C_WHITE,
     f"{len(events)} registros · Calendário completo"),
    ("C", "DB_ATIVIDADES", "✅ Banco de Dados\nde Atividades", "375623", C_WHITE,
     f"{len(activities)} atividades · Checklists consolidados"),
    ("D", "DB_FINANCEIRO", "💰 Banco de Dados\nFinanceiro", "843C0C", C_WHITE,
     f"{len(financeiro)} lançamentos · Budget e pagamentos"),
    ("E", "📊 DASHBOARD", "📊 Dashboard\nExecutivo", C_PURPLE, C_WHITE,
     "KPIs · Gráficos · Alertas"),
]
for col_l, _, label, bg, fg, sub in nav:
    r_start = 9
    ws.row_dimensions[r_start].height = 50
    ws.row_dimensions[r_start+1].height = 22
    ws.row_dimensions[r_start+2].height = 18
    col = ord(col_l) - ord('A') + 1
    ws.merge_cells(f"{col_l}{r_start}:{col_l}{r_start}")
    cell = ws.cell(row=r_start, column=col)
    cell.value = label; cell.fill = fx(bg)
    cell.font = ft(True, fg, 12); cell.alignment = al()
    cell.border = bdM()
    sub_cell = ws.cell(row=r_start+1, column=col)
    sub_cell.value = sub; sub_cell.fill = fx(C_LGRAY)
    sub_cell.font = ft(False, C_DGRAY if 'C_DGRAY' in dir() else "404040", 9)
    sub_cell.alignment = al(); sub_cell.border = bd()

# Summary stats row
ws.row_dimensions[13].height = 18
ws.merge_cells('B13:H13')
c = ws['B13']
evs_2026 = sum(1 for e in events if e['Ano']=='2026')
pend = sum(1 for a in activities if a['Status']=='PENDENTE')
total_prev = sum(f['Orçamento Previsto'] for f in financeiro)
c.value = (f"  Resumo  ·  Total Eventos: {len(events)}  ·  Eventos 2026: {evs_2026}  ·  "
           f"Atividades Pendentes: {pend}  ·  Budget Total: R$ {total_prev:,.0f}")
c.fill = fx(C_LBLUE); c.font = ft(True, C_NAVY, 10); c.alignment = al("left")

# ═════════════════════════════════════════════════════════════════════════════
# 2. DB_EVENTOS
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_EVENTOS")
ws.sheet_view.showGridLines = False
freeze(ws, "C4")

# Title
ws.merge_cells('A1:V1')
c=ws['A1']; c.value="📋 BANCO DE DADOS CENTRAL DE EVENTOS"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,14); c.alignment=al()
ws.row_dimensions[1].height=30

ws.merge_cells('A2:V2')
c=ws['A2']; c.value="Fonte única de verdade · Todos os eventos em uma tabela estruturada · Não edite linhas de outros eventos sem autorização"
c.fill=fx(C_LBLUE); c.font=ft(False,C_NAVY,9); c.alignment=al("left")
ws.row_dimensions[2].height=16

COLS_EV = [
    ('ID','A',4),('Evento','B',42),('Tipo','C',12),('EM','D',16),
    ('Status Evento','E',14),('Status EM','F',14),
    ('Ano','G',6),('Data Início','H',13),('Data Fim','I',13),
    ('Trimestre','J',10),('BU','K',16),('Produto','L',18),
    ('Solicitante','M',18),('Cidade','N',16),('Formato','O',13),
    ('Local','P',30),('Complexidade','Q',14),('Agência','R',16),
    ('Hotel','S',20),('Sociedade','T',30),('Stand','U',10),
    ('PSs Plan.','V',9),('PSs Pres.','W',9),('Observações','X',40),
]
for i,(label,col_l,w) in enumerate(COLS_EV, 1):
    ws.column_dimensions[col_l].width = w

ws.row_dimensions[3].height = 32
hdr(ws, 3, range(1, len(COLS_EV)+1))
for i,(label,_,_) in enumerate(COLS_EV,1):
    ws.cell(row=3, column=i).value = label

# Add table row filter
from openpyxl.worksheet.table import Table, TableStyleInfo
tab = Table(displayName="TblEventos", ref=f"A3:{get_column_letter(len(COLS_EV))}{3+len(events)}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,
    showLastColumn=False,showRowStripes=True,showColumnStripes=False)
ws.add_table(tab)

today = datetime.date.today()
def quarter(d):
    if d is None: return ''
    try: return f"Q{(d.month-1)//3+1}"
    except: return ''

for i, ev in enumerate(events):
    r = 4 + i
    data_row(ws, r, len(COLS_EV), alt=(i%2==1))
    row_data = [
        i+1, ev['Evento'], TIPO_MAP.get(ev['Tipo'], ev['Tipo']), ev['EM'],
        ev['Status Evento'], ev['Status EM'],
        ev['Ano'], ev['Data Início'], ev['Data Fim'],
        quarter(ev['Data Início']),
        ev['BU'], ev['Produto'], ev['Solicitante'], ev['Cidade'],
        ev['Formato'], ev['Local'], ev['Complexidade'], ev['Agência'],
        ev['Hotel'], ev['Sociedade'], ev['Stand'],
        ev['PSs Planejados'], ev['PSs Presentes'], ev['Observações']
    ]
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c_idx)
        cell.value = val
        if c_idx in (8,9) and isinstance(val, datetime.date):
            cell.number_format = 'DD/MM/YYYY'
        # Color status columns
        if c_idx == 5:
            status_cell(cell, ev['Status Evento'])
        if c_idx == 6:
            status_cell(cell, ev['Status EM'])

print("DB_EVENTOS criado")

# ═════════════════════════════════════════════════════════════════════════════
# 3. DB_ATIVIDADES
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_ATIVIDADES")
ws.sheet_view.showGridLines = False
freeze(ws, "C4")

ws.merge_cells('A1:H1')
c=ws['A1']; c.value="✅ BANCO DE DADOS DE ATIVIDADES"
c.fill=fx("375623"); c.font=ft(True,C_WHITE,14); c.alignment=al()
ws.row_dimensions[1].height=30

ws.merge_cells('A2:H2')
c=ws['A2']; c.value="Checklists de todos os eventos consolidados · Filtrar por Evento, Status ou Responsável"
c.fill=fx(C_LGREEN); c.font=ft(False,"375623",9); c.alignment=al("left")
ws.row_dimensions[2].height=16

COLS_AT = [
    ('ID','A',5),('Evento','B',38),('Etapa','C',20),('Atividade','D',45),
    ('Responsável','E',18),('% Conclusão','F',13),('Status','G',16),('Observações','H',40),
]
for label,col_l,w in COLS_AT:
    ws.column_dimensions[col_l].width = w

ws.row_dimensions[3].height = 28
hdr(ws, 3, range(1, len(COLS_AT)+1), bg="375623")
for i,(label,_,_) in enumerate(COLS_AT,1):
    ws.cell(row=3, column=i).value = label

tab2 = Table(displayName="TblAtividades", ref=f"A3:{get_column_letter(len(COLS_AT))}{3+len(activities)}")
tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7",showFirstColumn=False,
    showLastColumn=False,showRowStripes=True,showColumnStripes=False)
ws.add_table(tab2)

for i, at in enumerate(activities):
    r = 4 + i
    data_row(ws, r, len(COLS_AT), alt=(i%2==1))
    pct_val = at['% Conclusão']
    row_data = [i+1, at['Evento'], at['Etapa'], at['Atividade'],
                at['Responsável'], pct_val/100 if pct_val else 0,
                at['Status'], at['Observações']]
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c_idx)
        cell.value = val
        if c_idx == 6:
            cell.number_format = '0%'
            cell.alignment = al()
            # Color bar by percentage
            if pct_val == 100: cell.fill = fx("E2EFDA"); cell.font = ft(True,"375623",9)
            elif pct_val >= 50: cell.fill = fx("FFF2CC"); cell.font = ft(True,"7F6000",9)
            elif pct_val > 0:   cell.fill = fx("FCE4D6"); cell.font = ft(True,"843C0C",9)
            else:               cell.fill = fx("FFDDD8"); cell.font = ft(True,"C00000",9)
        if c_idx == 7:
            status_cell(cell, at['Status'])

# Data bar conditional formatting on % column (F4:F...)
last_r = 3 + len(activities)
ws.conditional_formatting.add(f"F4:F{last_r}",
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color="2E75B6"))

print("DB_ATIVIDADES criado")

# ═════════════════════════════════════════════════════════════════════════════
# 4. DB_FINANCEIRO
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_FINANCEIRO")
ws.sheet_view.showGridLines = False
freeze(ws, "C4")

ws.merge_cells('A1:M1')
c=ws['A1']; c.value="💰 BANCO DE DADOS FINANCEIRO"
c.fill=fx("843C0C"); c.font=ft(True,C_WHITE,14); c.alignment=al()
ws.row_dimensions[1].height=30

ws.merge_cells('A2:M2')
c=ws['A2']; c.value="Controle de budget, POs, pagamentos e NFs · Valores em R$"
c.fill=fx("FCE4D6"); c.font=ft(False,"843C0C",9); c.alignment=al("left")
ws.row_dimensions[2].height=16

COLS_FIN = [
    ('ID','A',5),('Evento','B',36),('EM','C',14),('FY','D',6),
    ('Fornecedor','E',22),('Tipo de Custo','F',20),('PO/WF','G',14),
    ('Orç. Previsto','H',16),('Real. FY2025','I',14),('Real. FY2026','J',14),
    ('Total Evento','K',16),('Status Pag.','L',14),('Status NF/ND','M',14),('Observações','N',38),
]
for label,col_l,w in COLS_FIN:
    ws.column_dimensions[col_l].width = w

ws.row_dimensions[3].height = 28
hdr(ws, 3, range(1, len(COLS_FIN)+1), bg="843C0C")
for i,(label,_,_) in enumerate(COLS_FIN,1):
    ws.cell(row=3, column=i).value = label

tab3 = Table(displayName="TblFinanceiro", ref=f"A3:{get_column_letter(len(COLS_FIN))}{3+len(financeiro)}")
tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3",showFirstColumn=False,
    showLastColumn=False,showRowStripes=True,showColumnStripes=False)
ws.add_table(tab3)

total_prev_all = sum(f['Orçamento Previsto'] for f in financeiro)
total_r25 = sum(f['Realizado FY2025'] for f in financeiro)
total_r26 = sum(f['Realizado FY2026'] for f in financeiro)

for i, fin in enumerate(financeiro):
    r = 4 + i
    data_row(ws, r, len(COLS_FIN), alt=(i%2==1))
    row_data = [i+1, fin['Evento'], fin['EM'], fin['FY'], fin['Fornecedor'],
                fin['Tipo de Custo'], fin['PO/WF'],
                fin['Orçamento Previsto'], fin['Realizado FY2025'],
                fin['Realizado FY2026'], fin['Total Evento'],
                fin['Status Pagamento'], fin['Status NF/ND'], fin['Observações']]
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c_idx)
        cell.value = val
        if c_idx in (8,9,10,11) and isinstance(val,(int,float)):
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = al("right")
        if c_idx == 12: status_cell(cell, fin['Status Pagamento'])
        if c_idx == 13: status_cell(cell, fin['Status NF/ND'])

# Totals row
r_tot = 4 + len(financeiro)
ws.row_dimensions[r_tot].height = 22
ws.merge_cells(f'A{r_tot}:G{r_tot}')
c = ws.cell(row=r_tot, column=1)
c.value = "TOTAIS GERAIS"; c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,10); c.alignment=al()
for col_idx, total in [(8, total_prev_all),(9, total_r25),(10, total_r26),(11, total_r25+total_r26)]:
    cell = ws.cell(row=r_tot, column=col_idx)
    cell.value = total; cell.number_format = 'R$ #,##0.00'
    cell.fill = fx(C_NAVY); cell.font = ft(True, C_WHITE, 10)
    cell.alignment = al("right"); cell.border = bd()

print("DB_FINANCEIRO criado")

# ═════════════════════════════════════════════════════════════════════════════
# 5. DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("📊 DASHBOARD")
ws.sheet_view.showGridLines = False

# Setup columns
for c in range(1, 16):
    ws.column_dimensions[get_column_letter(c)].width = 14

ws.column_dimensions['A'].width = 2

# Title
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 40
ws.merge_cells('B2:O2')
c=ws['B2']; c.value="📊 DASHBOARD EXECUTIVO · GESTÃO DE EVENTOS"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,20); c.alignment=al()

ws.merge_cells('B3:O3')
c=ws['B3']; c.value=f"Atualizado em: {datetime.date.today().strftime('%d/%m/%Y')}  ·  Fonte: DB_EVENTOS / DB_ATIVIDADES / DB_FINANCEIRO"
c.fill=fx(C_LBLUE); c.font=ft(False,C_NAVY,10); c.alignment=al("left")
ws.row_dimensions[3].height=18

# ── KPI Cards row ─────────────────────────────────────────────────────────────
ws.row_dimensions[5].height = 18  # label
ws.row_dimensions[6].height = 50  # value
ws.row_dimensions[7].height = 22  # sub

kpis = []
# Total events
kpis.append(("TOTAL\nEVENTOS", len(events), "", C_NAVY, C_WHITE))
# Events 2026
ev_2026 = [e for e in events if e['Ano']=='2026']
kpis.append(("EVENTOS\n2026", len(ev_2026), "", C_BLUE, C_WHITE))
# To realize
a_realizar = sum(1 for e in events if 'Realizar' in e['Status Evento'])
kpis.append(("A\nREALIZAR", a_realizar, "", C_ORANGE, C_WHITE))
# Realized
realizados = sum(1 for e in events if e['Status Evento']=='Realizado')
kpis.append(("REALIZADOS", realizados, "", C_GREEN, C_WHITE))
# Pending activities
pend_at = sum(1 for a in activities if a['Status']=='PENDENTE')
kpis.append(("ATIVIDADES\nPENDENTES", pend_at, "", "843C0C", C_WHITE))
# Budget
total_b = sum(f['Orçamento Previsto'] for f in financeiro)
kpis.append(("BUDGET\nTOTAL", f"R$ {total_b/1e6:.1f}M", "", C_PURPLE, C_WHITE))

kpi_cols = [2,4,6,8,10,12]
for idx, (label, value, sub, bg, fg) in enumerate(kpis):
    col = kpi_cols[idx]
    col2 = col+1
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col2)
    ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col2)
    ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col2)
    c=ws.cell(row=5,column=col); c.value=label; c.fill=fx(bg)
    c.font=ft(True,fg,9); c.alignment=al(); c.border=bd()
    c=ws.cell(row=6,column=col); c.value=value; c.fill=fx(bg)
    c.font=ft(True,fg,28); c.alignment=al(); c.border=bd()
    c=ws.cell(row=7,column=col); c.value=sub; c.fill=fx(C_LGRAY)
    c.font=ft(False,"404040",8); c.alignment=al(); c.border=bd()

# ── Status breakdown table ────────────────────────────────────────────────────
ws.row_dimensions[9].height = 22
ws.row_dimensions[10].height = 18

from collections import Counter
status_count = Counter(e['Status Evento'] for e in events)
bu_count = Counter(e['BU'] for e in events if e['BU'])
city_count = Counter(e['Cidade'] for e in events if e['Cidade'])
q_count = Counter()
for e in events:
    if e['Data Início'] and e['Ano']=='2026':
        q = f"Q{(e['Data Início'].month-1)//3+1} 2026"
        q_count[q] += 1

# Status por Evento table (B9)
ws.merge_cells('B9:E9')
c=ws['B9']; c.value="STATUS DOS EVENTOS"; c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,11); c.alignment=al()

headers_st = ['Status','Qtd','%']
for idx, h in enumerate(headers_st, 2):
    c=ws.cell(row=10, column=idx); c.value=h
    c.fill=fx(C_BLUE); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()

r=11
for stat, cnt in sorted(status_count.items(), key=lambda x:-x[1]):
    pct = cnt/len(events)*100 if events else 0
    for ci, val in enumerate([stat, cnt, f"{pct:.0f}%"], 2):
        c = ws.cell(row=r, column=ci); c.value=val
        c.font=ft(size=9); c.alignment=al(); c.border=bd()
        if ci == 2: status_cell(c, stat)
        else: c.fill = fx(C_LGRAY if r%2==1 else C_WHITE)
    r+=1

# BU Breakdown (G9)
ws.merge_cells('G9:J9')
c=ws['G9']; c.value="EVENTOS POR BU"; c.fill=fx("375623"); c.font=ft(True,C_WHITE,11); c.alignment=al()
for idx, h in enumerate(['BU','Qtd'], 7):
    c=ws.cell(row=10,column=idx); c.value=h
    c.fill=fx("375623"); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()
r=11
for bu, cnt in sorted(bu_count.items(), key=lambda x:-x[1])[:8]:
    for ci, val in enumerate([bu,cnt],7):
        c=ws.cell(row=r,column=ci); c.value=val
        c.fill=fx(C_LGRAY if r%2==1 else C_WHITE); c.font=ft(size=9)
        c.alignment=al(); c.border=bd()
    r+=1

# City top (K9)
ws.merge_cells('K9:N9')
c=ws['K9']; c.value="TOP CIDADES"; c.fill=fx(C_PURPLE); c.font=ft(True,C_WHITE,11); c.alignment=al()
for idx, h in enumerate(['Cidade','Qtd'],11):
    c=ws.cell(row=10,column=idx); c.value=h
    c.fill=fx(C_PURPLE); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()
r=11
for city, cnt in sorted(city_count.items(), key=lambda x:-x[1])[:8]:
    for ci, val in enumerate([city,cnt],11):
        c=ws.cell(row=r,column=ci); c.value=val
        c.fill=fx(C_LGRAY if r%2==1 else C_WHITE); c.font=ft(size=9)
        c.alignment=al(); c.border=bd()
    r+=1

# ── Activities status breakdown ───────────────────────────────────────────────
r_base = 22
ws.merge_cells(f'B{r_base}:E{r_base}')
c=ws.cell(row=r_base,column=2); c.value="ATIVIDADES POR STATUS"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,11); c.alignment=al()

act_status = Counter(a['Status'] for a in activities)
for idx,h in enumerate(['Status','Qtd','%'],2):
    c=ws.cell(row=r_base+1,column=idx); c.value=h
    c.fill=fx(C_BLUE); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()

r = r_base+2
for stat,cnt in sorted(act_status.items(), key=lambda x:-x[1]):
    pct = cnt/len(activities)*100
    for ci,val in enumerate([stat,cnt,f"{pct:.0f}%"],2):
        c=ws.cell(row=r,column=ci); c.value=val
        c.font=ft(size=9); c.alignment=al(); c.border=bd()
        if ci==2: status_cell(c,stat)
        else: c.fill=fx(C_LGRAY if r%2==0 else C_WHITE)
    r+=1

# ── Financial summary ─────────────────────────────────────────────────────────
ws.merge_cells(f'G{r_base}:J{r_base}')
c=ws.cell(row=r_base,column=7); c.value="RESUMO FINANCEIRO POR EVENTO"
c.fill=fx("843C0C"); c.font=ft(True,C_WHITE,11); c.alignment=al()

# Group financial by event
from collections import defaultdict
fin_by_ev = defaultdict(lambda:{'prev':0,'r25':0,'r26':0})
for f in financeiro:
    ev_k = f['Evento']
    fin_by_ev[ev_k]['prev'] += f['Orçamento Previsto']
    fin_by_ev[ev_k]['r25']  += f['Realizado FY2025']
    fin_by_ev[ev_k]['r26']  += f['Realizado FY2026']

for idx,h in enumerate(['Evento','Previsto','Real.Total'],7):
    c=ws.cell(row=r_base+1,column=idx); c.value=h
    c.fill=fx("843C0C"); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()

r=r_base+2
for ev_k, vals in sorted(fin_by_ev.items(), key=lambda x:-x[1]['prev'])[:10]:
    real_tot = vals['r25']+vals['r26']
    for ci,val in enumerate([ev_k[:30], vals['prev'], real_tot],7):
        c=ws.cell(row=r,column=ci); c.value=val
        if ci in (8,9): c.number_format='R$ #,##0'; c.alignment=al("right")
        else: c.alignment=al("left")
        c.fill=fx(C_LGRAY if r%2==0 else C_WHITE); c.font=ft(size=9); c.border=bd()
    r+=1

# Totals
for ci,val in enumerate(['TOTAL', total_prev_all, total_r25+total_r26],7):
    c=ws.cell(row=r,column=ci); c.value=val
    c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,10)
    if ci in (8,9): c.number_format='R$ #,##0'; c.alignment=al("right")
    else: c.alignment=al()
    c.border=bd()

# ── Payment status ─────────────────────────────────────────────────────────────
ws.merge_cells(f'K{r_base}:N{r_base}')
c=ws.cell(row=r_base,column=11); c.value="STATUS DE PAGAMENTOS"
c.fill=fx(C_PURPLE); c.font=ft(True,C_WHITE,11); c.alignment=al()

pay_status = Counter(f['Status Pagamento'] for f in financeiro if f['Status Pagamento'])
for idx,h in enumerate(['Status','Qtd'],11):
    c=ws.cell(row=r_base+1,column=idx); c.value=h
    c.fill=fx(C_PURPLE); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()
r=r_base+2
for stat,cnt in sorted(pay_status.items(), key=lambda x:-x[1]):
    for ci,val in enumerate([stat,cnt],11):
        c=ws.cell(row=r,column=ci); c.value=val
        c.font=ft(size=9); c.alignment=al(); c.border=bd()
        if ci==11: status_cell(c,stat)
        else: c.fill=fx(C_LGRAY if r%2==0 else C_WHITE)
    r+=1

print("DASHBOARD criado")

# ═════════════════════════════════════════════════════════════════════════════
# 6. PARTICIPAÇÃO
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("PARTICIPAÇÃO")
ws.sheet_view.showGridLines = False
freeze(ws, "B4")

ws.merge_cells('A1:K1')
c=ws['A1']; c.value="✈️ CONTROLE DE PARTICIPAÇÃO E LOGÍSTICA"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,14); c.alignment=al()
ws.row_dimensions[1].height=30

ws.merge_cells('A2:K2')
c=ws['A2']; c.value="Registro de deslocamentos, hospedagem e custos por participante"
c.fill=fx(C_LBLUE); c.font=ft(False,C_NAVY,9); c.alignment=al("left")
ws.row_dimensions[2].height=16

COLS_PART = [
    ('Evento','A',35),('Participante','B',22),('Cargo','C',18),('Regional','D',10),
    ('Cidade Origem','E',15),('Data Ida','F',12),('Data Volta','G',12),
    ('Aéreo R$','H',12),('Hospedagem R$','I',14),('Refeição R$','J',12),
    ('Deslocamento R$','K',16),('Total R$','L',12),('Observações','M',30),
]
for label,col_l,w in COLS_PART:
    ws.column_dimensions[col_l].width = w

ws.row_dimensions[3].height=28
hdr(ws,3,range(1,len(COLS_PART)+1),bg=C_NAVY)
for i,(label,_,_) in enumerate(COLS_PART,1):
    ws.cell(row=3,column=i).value=label

# Copy existing participation data
part_ws = src['PARTICIPAÇÃO']
r=4
for row in part_ws.iter_rows(min_row=3,max_row=part_ws.max_row,values_only=True):
    if row[1] is None: continue
    data_row(ws,r,len(COLS_PART),alt=(r%2==0))
    vals = [row[1],None,None,None,None]
    if isinstance(row[4],datetime.datetime): vals.append(row[4].date())
    else: vals.append(row[4])
    if isinstance(row[5],datetime.datetime): vals.append(row[5].date())
    else: vals.append(row[5])
    try: aer=float(str(row[6])) if row[6] else 0
    except: aer=0
    try: hosp=float(str(row[7])) if row[7] else 0
    except: hosp=0
    try: ref=float(str(row[8])) if row[8] else 0
    except: ref=0
    try: desl=float(str(row[9])) if row[9] else 0
    except: desl=0
    vals += [aer,hosp,ref,desl,aer+hosp+ref+desl,'']
    for ci,val in enumerate(vals,1):
        cell=ws.cell(row=r,column=ci); cell.value=val
        if ci in (8,9,10,11,12) and isinstance(val,(int,float)):
            cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
    r+=1

tab4 = Table(displayName="TblParticipacao", ref=f"A3:{get_column_letter(len(COLS_PART))}{r-1}")
tab4.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tab4)

print("PARTICIPAÇÃO criado")

# ═════════════════════════════════════════════════════════════════════════════
# 7. ROADMAP
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🗺️ ROADMAP")
ws.sheet_view.showGridLines = False
for c in range(1,10):
    ws.column_dimensions[get_column_letter(c)].width = 22
ws.column_dimensions['A'].width = 3

ws.row_dimensions[1].height=10
ws.row_dimensions[2].height=40
ws.merge_cells('B2:I2')
c=ws['B2']; c.value="🗺️ ROADMAP DE IMPLEMENTAÇÃO"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,20); c.alignment=al()

phases = [
    ("FASE 1 — QUICK WINS", C_GREEN, "E2EFDA", "375623", [
        ("✅ Concluído","Migração de dados para DB_EVENTOS centralizado"),
        ("✅ Concluído","Consolidação de checklists em DB_ATIVIDADES"),
        ("✅ Concluído","Unificação financeira em DB_FINANCEIRO"),
        ("✅ Concluído","Formatação de tabelas com Excel Tables (filtros)"),
        ("✅ Concluído","Dashboard executivo com KPIs e resumos"),
        ("✅ Concluído","Padronização de status com formatação condicional"),
        ("🔄 Próximo","Treinar equipe no uso dos novos bancos de dados"),
    ]),
    ("FASE 2 — REESTRUTURAÇÃO", C_BLUE, "BDD7EE", "1F3864", [
        ("🔄 Próximo","Criar tabela de Fornecedores cadastrados (DB_Fornecedores)"),
        ("🔄 Próximo","Criar tabela de Participantes (DB_Participantes)"),
        ("🔄 Próximo","Implementar validação de dados em campos de status"),
        ("🔄 Próximo","Padronizar nomenclatura de eventos (chave única por EM)"),
        ("🔄 Próximo","Criar modelo padrão de checklist para novos eventos"),
        ("🔄 Próximo","Revisão e validação de todos os dados migrados"),
    ]),
    ("FASE 3 — AUTOMAÇÃO", C_ORANGE, "FCE4D6", "843C0C", [
        ("📅 Futuro","Power Query: atualização automática de relatórios"),
        ("📅 Futuro","Office Scripts: geração de relatório semanal por e-mail"),
        ("📅 Futuro","Alertas automáticos de prazos via conditional formatting"),
        ("📅 Futuro","Power Automate: notificação de pagamentos pendentes"),
        ("📅 Futuro","Power Automate: lembrete de atividades atrasadas"),
        ("📅 Futuro","Integração com SharePoint para controle de versão"),
    ]),
    ("FASE 4 — DASHBOARD AVANÇADO", C_PURPLE, "EAD1FF", "7030A0", [
        ("📅 Futuro","Power BI: dashboard interativo com drill-down"),
        ("📅 Futuro","Gráfico de Gantt de eventos por trimestre"),
        ("📅 Futuro","Mapa geográfico de eventos por cidade"),
        ("📅 Futuro","Análise de tendências de budget YoY"),
        ("📅 Futuro","Relatório executivo automatizado (PDF)"),
        ("📅 Futuro","Scorecard de performance por responsável"),
    ]),
]

r = 4
for phase_title, bg_hdr, bg_light, fg_dark, items in phases:
    ws.row_dimensions[r].height=28
    ws.merge_cells(f'B{r}:I{r}')
    c=ws.cell(row=r,column=2); c.value=phase_title
    c.fill=fx(bg_hdr if len(bg_hdr)==6 else C_NAVY); c.font=ft(True,C_WHITE,12); c.alignment=al()
    r+=1
    
    for hcol,htxt in enumerate(['Status','Atividade'],2):
        c=ws.cell(row=r,column=hcol); c.value=htxt
        c.fill=fx(bg_light); c.font=ft(True,fg_dark,9); c.alignment=al(); c.border=bd()
    ws.row_dimensions[r].height=18; r+=1
    
    for stat,desc in items:
        ws.row_dimensions[r].height=20
        c=ws.cell(row=r,column=2); c.value=stat
        bg_s = "E2EFDA" if "Concluído" in stat else ("FFF2CC" if "Próximo" in stat else "F2F2F2")
        c.fill=fx(bg_s); c.font=ft(True,fg_dark,9); c.alignment=al(); c.border=bd()
        ws.merge_cells(f'C{r}:I{r}')
        c=ws.cell(row=r,column=3); c.value=desc
        c.fill=fx(C_WHITE if r%2==0 else C_LGRAY); c.font=ft(size=9)
        c.alignment=al("left"); c.border=bd()
        r+=1
    r+=1

print("ROADMAP criado")

# ═════════════════════════════════════════════════════════════════════════════
# Sheet tab colors
# ═════════════════════════════════════════════════════════════════════════════
tab_colors = {
    "🏠 INÍCIO":       "1F3864",
    "DB_EVENTOS":      "2E75B6",
    "DB_ATIVIDADES":   "375623",
    "DB_FINANCEIRO":   "843C0C",
    "📊 DASHBOARD":    "7030A0",
    "PARTICIPAÇÃO":    "2E75B6",
    "🗺️ ROADMAP":     "404040",
}
for sname, color in tab_colors.items():
    if sname in wb.sheetnames:
        wb[sname].sheet_properties.tabColor = color

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = '/home/user/Femartignon/NEC_Eventos_Reestruturado_v2.xlsx'
wb.save(out_path)
print(f"\n✅ Arquivo salvo em: {out_path}")
