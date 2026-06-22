"""
Preencher coluna Prazo (F) de DB_ATIVIDADES calculando a partir da Data Início
do evento, perfil "Mais folga" escolhido pelo usuário:
  Planejamento -90 | VEEVA -75 | Sociedade -60 | Agência -45 |
  Simpósio -30 | Stand -21 | Alinhamentos Finais -14 | Pós-Evento +15
Só preenche atividades PENDENTE e EM ANDAMENTO. Pula CONCLUÍDO.
Eventos sem data ou etapa não-classificada ficam em branco (para revisão).
"""
import datetime
from openpyxl import load_workbook

PATH = 'NEC_Eventos_Reestruturado_v4.xlsx'
wb = load_workbook(PATH)
ev = wb['DB_EVENTOS']
at = wb['DB_ATIVIDADES']

# ── 1. Mapa Evento -> menor Data Início (col D=4, K=11) ────────────────────────
ev_date = {}
for r in range(4, ev.max_row+1):
    nome = ev.cell(row=r, column=4).value
    dt = ev.cell(row=r, column=11).value
    if not nome or dt is None: continue
    if isinstance(dt, datetime.datetime): dt = dt.date()
    nome = str(nome).strip()
    if nome not in ev_date or dt < ev_date[nome]:
        ev_date[nome] = dt

# ── 2. Override: nomes usados em ATIVIDADES -> nome canônico em DB_EVENTOS ──────
OVERRIDE = {
    'ABENEPI 2026': 'ABENEPI - VIII Congresso Internacional & XXVIII Brasileiro da ABENEPI',
    'APRENDER CRIANÇA 2026': 'APRENDER CRIANÇA',
    'BRAIN 2026 - Simpósio': 'BRAIN 2026',
    'Board Review Albert Einstein': 'Board Review: XX Curso de Revisão em Hematologia e Hemoterapia',
    'Congresso Medicina do Sono': 'XXIII Congresso Paulista de Medicina do sono 2026',
    'FORUM GEDIIB BRASÍLIA': 'FORUM NACIONAL DE ACESSO GEDIIB BRASÍLIA',
}
def data_do_evento(nome):
    nome = str(nome).strip()
    if nome in ev_date: return ev_date[nome]
    if nome in OVERRIDE: return ev_date.get(OVERRIDE[nome])
    return None

# ── 3. Classificador de Etapa -> offset em dias (negativo=antes, positivo=depois)
def offset_for(etapa):
    e = (etapa or '').lower().strip()
    if not e: return None
    # Pós-Evento: exigir "evento" junto (senão 'pós' casa dentro de 'simPÓSio')
    if ('pós' in e or 'pos' in e) and 'evento' in e:       return +15   # Pós-Evento
    if 'planej' in e or e == 'plan':                       return -90   # Planejamento
    if 'veeva' in e:                                        return -75   # VEEVA
    if any(k in e for k in ['socied','contrapart','cota','patroc','bolsa']): return -60
    if any(k in e for k in ['agênc','agenc','logíst','logist','produç','rsvp']): return -45
    if 'simpós' in e or 'simpos' in e or 'palestr' in e:   return -30   # Simpósio/Palestrantes
    if 'alinhament' in e:                                   return -14   # Alinhamentos Finais
    if any(k in e for k in ['stand','buffet','staff','credenci','logomarc',
            'divulg','encarte','vídeo','video','realiz','pórtic','portic',
            'totem','toten','palco','expositor','feira']):  return -21   # Execução/Stand
    return None  # desconhecido/vazio -> não preenche

# ── 4. Preencher ──────────────────────────────────────────────────────────────
preenchidos = 0
sem_data = []
sem_etapa = []
pulados_concluido = 0
for r in range(4, at.max_row+1):
    evn = at.cell(row=r, column=2).value
    et  = at.cell(row=r, column=3).value
    st  = at.cell(row=r, column=8).value
    ativ = at.cell(row=r, column=4).value
    if evn is None: continue
    status = str(st).strip() if st else ''
    if status == 'CONCLUÍDO':
        pulados_concluido += 1
        continue
    if status not in ('PENDENTE', 'EM ANDAMENTO'):
        continue
    base = data_do_evento(evn)
    if base is None:
        sem_data.append((r, str(evn).strip(), str(ativ)[:35] if ativ else ''))
        continue
    off = offset_for(et)
    if off is None:
        sem_etapa.append((r, str(et).strip() if et else '(vazio)', str(ativ)[:35] if ativ else ''))
        continue
    prazo = base + datetime.timedelta(days=off)
    cell = at.cell(row=r, column=6)
    cell.value = prazo
    cell.number_format = 'DD/MM/YYYY'
    preenchidos += 1

wb.save(PATH)

print(f"✅ Prazos preenchidos: {preenchidos}")
print(f"   Pulados (CONCLUÍDO): {pulados_concluido}")
print(f"   Sem data de evento (em branco): {len(sem_data)}")
for r, e, a in sem_data:
    print(f"      L{r}: evento={repr(e)} | {a}")
print(f"   Etapa não classificada (em branco): {len(sem_etapa)}")
for r, e, a in sem_etapa:
    print(f"      L{r}: etapa={repr(e)} | {a}")

# ── Verificação: amostra de prazos preenchidos ────────────────────────────────
wb2 = load_workbook(PATH)
at2 = wb2['DB_ATIVIDADES']
print("\n=== Amostra de prazos preenchidos (PENDENTE/EM ANDAMENTO) ===")
shown = 0
for r in range(4, at2.max_row+1):
    st = at2.cell(row=r, column=8).value
    f = at2.cell(row=r, column=6).value
    if f and str(st).strip() in ('PENDENTE','EM ANDAMENTO'):
        evn = at2.cell(row=r, column=2).value
        et = at2.cell(row=r, column=3).value
        print(f"   L{r}: {str(evn)[:22]:22} | {str(et)[:14]:14} | prazo={f.strftime('%d/%m/%Y') if hasattr(f,'strftime') else f} | {st}")
        shown += 1
    if shown >= 12: break
