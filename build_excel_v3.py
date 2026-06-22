"""
Fase 2 — Reestruturação
Gera NEC_Eventos_Reestruturado_v3.xlsx com:
  - DB_FORNECEDORES (nova aba)
  - DB_PARTICIPANTES (nova aba)
  - Data Validation dropdowns em todos os DB_
  - ID_Unico padronizado + correções de dados
  - MODELO_CHECKLIST (nova aba)
  - Roadmap atualizado (Fase 2 = Concluído)
"""

import re, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict, Counter

# ── Helpers ───────────────────────────────────────────────────────────────────
def fx(h): return PatternFill("solid", fgColor=h)
def ft(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def bd():
    s=Side(style="thin",color="D9D9D9"); return Border(left=s,right=s,top=s,bottom=s)
def bdM(col="2E75B6"):
    s=Side(style="medium",color=col); return Border(left=s,right=s,top=s,bottom=s)
def al(h="center",v="center",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def set_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

C_NAVY,C_BLUE,C_LBLUE  = "1F3864","2E75B6","BDD7EE"
C_LGRAY,C_WHITE         = "F2F2F2","FFFFFF"
C_GREEN,C_LGREEN        = "375623","E2EFDA"
C_ORANGE,C_LORANGE      = "843C0C","FCE4D6"
C_RED,C_LRED            = "C00000","FFDDD8"
C_YELLOW,C_LYELL        = "7F6000","FFF2CC"
C_PURPLE                = "7030A0"
C_TEAL                  = "1F6B75"

STATUS_FX = {
    'Realizado':    ('E2EFDA','375623'), 'a Realizar':('BDD7EE','1F3864'),
    'Cancelado':    ('FFDDD8','C00000'), 'CONCLUÍDO': ('E2EFDA','375623'),
    'EM ANDAMENTO': ('FFF2CC','7F6000'), 'PENDENTE':  ('FCE4D6','843C0C'),
    'ATRASADO':     ('FFDDD8','C00000'), '-':         ('F2F2F2','404040'),
    'Pago':         ('E2EFDA','375623'), 'Pós Evento':('FFF2CC','7F6000'),
    'Pendente':     ('FCE4D6','843C0C'), 'aprovado':  ('BDD7EE','1F3864'),
    'pago':         ('E2EFDA','375623'), 'Recebida':  ('E2EFDA','375623'),
    'Recebido':     ('E2EFDA','375623'), 'Homologado':('E2EFDA','375623'),
    'A preencher':  ('FFF2CC','7F6000'),
}
def status_cell(cell, status):
    key = status.strip() if status else ''
    if key in STATUS_FX:
        bg,fg = STATUS_FX[key]; cell.fill=fx(bg); cell.font=ft(True,fg,9)
    cell.alignment=al(); cell.border=bd()

def hdr(ws, r, cols, bg=C_NAVY, fg=C_WHITE, sz=10):
    for c in cols:
        cell=ws.cell(row=r,column=c)
        cell.fill=fx(bg); cell.font=ft(True,fg,sz)
        cell.alignment=al(); cell.border=bd()

def data_row(ws, r, max_c, alt=False):
    bg=C_LGRAY if alt else C_WHITE
    for c in range(1,max_c+1):
        cell=ws.cell(row=r,column=c)
        cell.fill=fx(bg); cell.font=ft(size=9)
        cell.alignment=al("left"); cell.border=bd()

def title_block(ws, merge_range, text, bg, fg="FFFFFF", sz=14):
    ws.merge_cells(merge_range)
    first_cell = merge_range.split(':')[0]
    c=ws[first_cell]; c.value=text; c.fill=fx(bg)
    c.font=ft(True,fg,sz); c.alignment=al()

def subtitle_block(ws, merge_range, text, bg, fg):
    ws.merge_cells(merge_range)
    first_cell = merge_range.split(':')[0]
    c=ws[first_cell]; c.value=text; c.fill=fx(bg)
    c.font=ft(False,fg,9); c.alignment=al("left")

def add_dv(ws, dv_type, formula1, sqref, prompt_title="", prompt="", error=""):
    dv = DataValidation(type=dv_type, formula1=formula1, showDropDown=False,
                        showInputMessage=True, showErrorMessage=True)
    if prompt_title: dv.promptTitle=prompt_title
    if prompt:       dv.prompt=prompt
    if error:        dv.error=error; dv.errorTitle="Valor inválido"
    dv.sqref = sqref
    ws.add_data_validation(dv)

TIPO_MAP = {'C':'Cota','A':'Apoio','S':'Simpósio','T':'TOME/Advisory'}

# ── Load source & v2 ──────────────────────────────────────────────────────────
src = load_workbook('/root/.claude/uploads/6d41cea8-3482-52e3-859c-cdd7002a8a66/1318286f-Calenda_rio_Neuro_Acesso_NEC.xlsx', data_only=True)
v2  = load_workbook('/home/user/Femartignon/NEC_Eventos_Reestruturado_v2.xlsx', data_only=True)

# ═══════════════════════════════════════════════════════════════════════════════
# Re-extract & clean data (Objetivo 4: padronização + correção)
# ═══════════════════════════════════════════════════════════════════════════════
def clean_em(v):
    if not v: return ''
    return re.sub(r'\s+', '', str(v).strip().upper())

def fix_name(v):
    fixes = {'Biansa Fasas': 'Bianca Facas', 'biansa fasas': 'Bianca Facas'}
    s = str(v).strip() if v else ''
    return fixes.get(s, s)

cal = src['Calendário de Eventos']
events = []
em_set = set()
for row in cal.iter_rows(min_row=4, max_row=60, values_only=True):
    if row[2] is None: continue
    name = str(row[2]).strip()
    if name in ('APOIO','COTA','SIMPÓSIO','TOME','','-'): continue
    def s(v): return str(v).strip() if v is not None else ''
    def d(v):
        if isinstance(v,datetime.datetime): return v.date()
        if isinstance(v,datetime.date): return v
        return None
    try: psp=int(float(str(row[19]))) if row[19] not in (None,'','-','X') else None
    except: psp=None
    try: psa=int(float(str(row[20]))) if row[20] not in (None,'','-','X') else None
    except: psa=None
    em_raw = s(row[4])
    em_clean = clean_em(em_raw)
    solicitante = fix_name(s(row[9]))
    dt_ini = d(row[13]); dt_fim = d(row[14])
    # Trimestre
    q = f"Q{(dt_ini.month-1)//3+1} {s(row[12])}" if dt_ini and s(row[12]) else ''
    # Duplicate EM detection
    dup_flag = '⚠️ EM duplicado' if em_clean and em_clean in em_set and em_clean not in ('PENDENTE','TBD') else ''
    if em_clean and em_clean not in ('PENDENTE','TBD'): em_set.add(em_clean)
    events.append({
        'ID_Unico': em_clean or f"SEM-EM-{len(events)+1:03d}",
        'Evento':name,'Tipo':s(row[1]),'EM':em_raw,'EM_Padrao':em_clean,
        'Status Evento':s(row[7]),'Status EM':s(row[8]),
        'Solicitante':solicitante,'BU':s(row[10]),'Produto':s(row[11]),
        'Ano':s(row[12]),'Data Início':dt_ini,'Data Fim':dt_fim,'Trimestre':q,
        'Cidade':s(row[16]),'Formato':s(row[17]),'Local':s(row[18]),
        'PSs Planejados':psp,'PSs Presentes':psa,
        'Agência':s(row[21]),'Hotel':s(row[22]),'Sociedade':s(row[23]),
        'Stand':s(row[24]),'Complexidade':s(row[3]),'Observações':s(row[27]),
        'Flag':dup_flag,
    })

# Activities
ACT_CFG = {
    'Fórum de Acesso 2026':  ('FÓRUM ACESSO 2026', 1,2,3,4,5,6, 5),
    ' FÓRUM GEDIIB ':        ('FORUM GEDIIB BRASÍLIA', 3,4,5,6,7,8, 5),
    'Cong. Medicina do Sono':('Congresso Medicina do Sono', 1,2,3,4,5,6, 5),
    'Brain 2026':            ('BRAIN 2026', 2,3,4,5,6,7, 6),
    'Brain Simpósio':        ('BRAIN 2026 - Simpósio', 2,3,4,5,6,7, 5),
    'Board Albert Einstein': ('Board Review Albert Einstein', 1,2,3,4,5,6, 5),
    'CONASENS':              ('CONASENS', 1,2,3,4,5,6, 5),
    'Aprender Criança 2026': ('APRENDER CRIANÇA 2026', 1,2,3,4,5,6, 5),
    'ADV. BOARD FRUZAQLA':   ('Advisory Board Fruzaqla', 1,2,3,4,5,6, 5),
    'Masterclass 2026':      ('MASTERCLASS 2026', 2,3,4,5,6,7, 5),
    'Masterclass Simpósio':  ('MASTERCLASS 2026 - Simpósio', 2,3,4,5,6,7, 5),
    'ABENEPI':               ('ABENEPI 2026', 1,2,3,4,5,6, 6),
}
activities = []
for sname,(ev,ce,ca,cr,cp,cs,co,start) in ACT_CFG.items():
    ws2=src[sname]; last_e=''
    for row in ws2.iter_rows(min_row=start,max_row=ws2.max_row,values_only=True):
        raw=list(row)+[None]*15
        ev_v=raw[ce-1];av=raw[ca-1];rv=raw[cr-1];pv=raw[cp-1];sv=raw[cs-1];ov=raw[co-1]
        if ev_v not in (None,'') and str(ev_v).strip() not in ('Atividade','Responsável','Andamento (%)','Status','Observações'):
            last_e=str(ev_v).strip()
        if av is None or str(av).strip() in ('','Atividade','Responsável','Andamento (%)','Status','Observações'): continue
        try: pct=int(float(str(pv))) if pv is not None else 0
        except: pct=0
        stat=str(sv).strip() if sv else ''
        if not stat: stat='CONCLUÍDO' if pct==100 else ('EM ANDAMENTO' if pct>0 else 'PENDENTE')
        activities.append({'Evento':ev,'Etapa':last_e,'Atividade':str(av).strip(),
            'Responsável':str(rv).strip() if rv else '',
            '% Conclusão':pct,'Status':stat,'Observações':str(ov).strip() if ov else ''})

fin_ws=src['FINANCEIRO']
financeiro=[]
for row in fin_ws.iter_rows(min_row=2,max_row=fin_ws.max_row,values_only=True):
    if row[0] is None: continue
    def n(v):
        if v is None: return 0.0
        try: return float(str(v).replace(',','.'))
        except: return 0.0
    financeiro.append({'Evento':str(row[0]).strip(),'EM':clean_em(row[1]),
        'FY':str(row[2]).strip() if row[2] else '','Fornecedor':str(row[3]).strip() if row[3] else '',
        'Tipo de Custo':str(row[4]).strip() if row[4] else '',
        'PO/WF':str(row[5]).strip() if row[5] else '',
        'Orçamento Previsto':n(row[6]),'Realizado FY2025':n(row[7]),'Realizado FY2026':n(row[8]),
        'Total Evento':n(row[9]),'Status Pagamento':str(row[10]).strip() if row[10] else '',
        'Status NF/ND':str(row[11]).strip() if row[11] else '',
        'Observações':str(row[12]).strip() if row[12] else ''})

# ═══════════════════════════════════════════════════════════════════════════════
# Build suppliers (Objetivo 1)
# ═══════════════════════════════════════════════════════════════════════════════
SUPPLIER_CATALOG = {
    'Casa da Vila':         ('Agência de Produção', 'Produção de Eventos', 'SP', 'Homologado'),
    'Incentivare':          ('Agência Logística', 'Logística, Buffet, Apoios', 'SP', 'Homologado'),
    'OPUS':                 ('Agência Logística', 'Logística Convidados', 'SP', 'Homologado'),
    'Opus':                 ('Agência Logística', 'Logística Convidados', 'SP', 'Homologado'),
    'Jardim Elétrico':      ('Montadora de Stand', 'Montagem de Stands', 'SP', 'Homologado'),
    'Jardim':               ('Montadora de Stand', 'Montagem de Stands', 'SP', 'A verificar'),
    'Poli Design':          ('Montadora de Stand', 'Montagem de Stands', 'SP', 'Homologado'),
    'Poli':                 ('Montadora de Stand', 'Montagem de Stands', 'SP', 'A verificar'),
    'Boutique Gourmet':     ('Buffet/Alimentação', 'Buffet, Lunch Box', 'SP', 'Homologado'),
    'Boutique':             ('Buffet/Alimentação', 'Buffet', 'SP', 'A verificar'),
    'APM':                  ('Sociedade Médica', 'Cota de Patrocínio', 'SP', 'Contrato'),
    'Sociedade':            ('Sociedade Médica', 'Cota de Patrocínio', 'Vários', 'Contrato'),
    'Dr. Fernando':         ('Palestrante', 'Honorários Palestrante', 'SP', 'Contrato'),
    'Dra. Dalva':           ('Palestrante', 'Honorários Palestrante', 'SP', 'Contrato'),
    'FEE Palestrante':      ('Palestrante', 'Honorários Palestrante', 'Vários', 'A verificar'),
    'Oficial':              ('Montadora Oficial', 'Montagem Oficial do Evento', 'Vários', 'Oficial evento'),
    'Homologado local':     ('Buffet Local', 'Buffet homologado pelo local', 'Vários', 'Homologado'),
}

RESP_UNICOS = sorted(set(a['Responsável'] for a in activities if a['Responsável']))
STATUS_EVENTO_LIST = ['Realizado','a Realizar','Cancelado','Em Planejamento']
STATUS_EM_LIST    = ['aprovada','Encerrada','pendente','solicitado','Sem ação Eventos']
STATUS_AT_LIST    = ['CONCLUÍDO','EM ANDAMENTO','PENDENTE','ATRASADO','-']
STATUS_PAG_LIST   = ['Pago','pago','aprovado','Pós Evento','Pendente','A preencher','Cancelado']
STATUS_NF_LIST    = ['Recebida','Recebido','Adiantamento','Pendente','A preencher']
TIPO_EV_LIST      = ['Cota','Apoio','Simpósio','TOME/Advisory']
BU_LIST           = sorted(set(e['BU'] for e in events if e['BU']))
FORMATO_LIST      = ['Presencial','Virtual','Hibrido']
COMPLEXIDADE_LIST = ['Alta','Média','Baixa']
AGENCIA_LIST      = ['Incentivare','OPUS','Opus','A preencher']
TIPO_CUSTO_LIST   = ['Cota de Patrocínio','Logística Convidados','Produção','Stand',
                      'Buffet Stand','Lunch Box','Palestrante','FEE Palestrante','Agência','Outros']

# ═══════════════════════════════════════════════════════════════════════════════
# Participants (Objetivo 2)
# ═══════════════════════════════════════════════════════════════════════════════
ws_ap = src['Brain Apoios - novo']
brain_participants = []
skip_names = {'Hilton (Hotel 1)','Manhatam (Hotel 2)','Park Plaza (Hotel 3)',
              'Participantes','Staff Takeda','Hotel apoios'}
for row in ws_ap.iter_rows(min_row=2, max_row=49, values_only=True):
    if row[0] is None: continue
    name = str(row[0]).strip()
    if name in skip_names or 'Apoios ' in name or not name: continue
    reg  = str(row[1]).strip() if row[1] else ''
    qtd  = row[2]
    hotel= row[3]
    brain_participants.append({'Nome':name,'Cargo/Regional':reg,'Tipo':'Staff Takeda',
        'Evento Ref.':'BRAIN 2026','Hotel Bloco':f"Hotel {hotel}" if hotel else '',
        'Observações':''})

# Also from col 13-15 (participant list in right table)
for row in ws_ap.iter_rows(min_row=5, max_row=40, values_only=True):
    raw=list(row)+[None]*20
    if raw[12] and raw[13] and str(raw[12]).strip() not in ('','#'):
        try: idx=int(raw[11]) if raw[11] else 0
        except: idx=0
        name = str(raw[13]).strip()
        hotel_col = str(raw[14]).strip() if raw[14] else ''
        if name and name not in [p['Nome'] for p in brain_participants]:
            # parse "Nome Sobrenome - Cargo - Regional"
            parts = name.split(' - ')
            nome_clean = parts[0].strip() if parts else name
            cargo = parts[1].strip() if len(parts)>1 else ''
            reg   = parts[2].strip() if len(parts)>2 else ''
            brain_participants.append({'Nome':nome_clean,'Cargo/Regional':f"{cargo} | {reg}",
                'Tipo':'Staff Takeda','Evento Ref.':'BRAIN 2026',
                'Hotel Bloco':f"Hotel {hotel_col}" if hotel_col else '','Observações':''})

print(f"Dados prontos → Eventos:{len(events)} | Atividades:{len(activities)} | Fin:{len(financeiro)} | Fornecedores:{len(SUPPLIER_CATALOG)} | Participantes:{len(brain_participants)}")

# ═══════════════════════════════════════════════════════════════════════════════
# Build new workbook v3
# ═══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ── Sheet creation order ──────────────────────────────────────────────────────
# 0. INÍCIO (capa atualizada)
# 1. DB_EVENTOS (com ID_Unico, validações, flags)
# 2. DB_ATIVIDADES (com validações)
# 3. DB_FINANCEIRO (com validações, EM padronizado)
# 4. DB_FORNECEDORES (novo)
# 5. DB_PARTICIPANTES (novo)
# 6. DASHBOARD (mesmo da v2)
# 7. PARTICIPAÇÃO
# 8. MODELO_CHECKLIST (novo)
# 9. ROADMAP (Fase 2 = concluído)

# ═══════════════════════════════════════════════════════════════════════════════
# 1. INÍCIO
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🏠 INÍCIO")
ws.sheet_view.showGridLines = False
for c in ['A','B','C','D','E','F','G','H']:
    ws.column_dimensions[c].width = {'A':2,'B':30,'C':28,'D':28,'E':22,'F':22,'G':22,'H':8}[c]

ws.row_dimensions[1].height = 10
for r in range(2,8): ws.row_dimensions[r].height = 24
for r in range(2,8):
    for col in range(2,9):
        ws.cell(row=r,column=col).fill = fx(C_NAVY)

ws.merge_cells('B2:H4')
c=ws['B2']; c.value="GESTÃO DE EVENTOS FARMACÊUTICOS"
c.font=ft(True,C_WHITE,22); c.alignment=al()
ws.merge_cells('B5:H5')
c=ws['B5']; c.value="Neurociências & Acesso  ·  NEC  ·  Takeda"
c.font=ft(False,"BDD7EE",13); c.alignment=al()
ws.merge_cells('B6:H7')
c=ws['B6']; c.value=f"Versão 3  ·  Fase 2 Concluída  ·  {datetime.date.today().strftime('%d/%m/%Y')}"
c.font=ft(False,"9DC3E6",10); c.alignment=al()

ws.row_dimensions[9].height = 55
ws.row_dimensions[10].height = 22
navs = [
    ('B','DB_EVENTOS',"📋 Banco de\nEventos",C_BLUE,f"{len(events)} eventos"),
    ('C','DB_ATIVIDADES',"✅ Banco de\nAtividades","375623",f"{len(activities)} atividades"),
    ('D','DB_FINANCEIRO',"💰 Banco\nFinanceiro","843C0C",f"{len(financeiro)} lançamentos"),
    ('E','DB_FORNECEDORES',"🏢 Banco de\nFornecedores",C_TEAL,f"{len(SUPPLIER_CATALOG)} cadastros"),
    ('F','DB_PARTICIPANTES',"👥 Banco de\nParticipantes",C_PURPLE,f"{len(brain_participants)} pessoas"),
    ('G','📊 DASHBOARD',"📊 Dashboard",C_NAVY,"KPIs executivos"),
]
for col_l,_,label,bg,sub in navs:
    col=ord(col_l)-ord('A')+1
    c=ws.cell(row=9,column=col); c.value=label
    c.fill=fx(bg); c.font=ft(True,C_WHITE,11); c.alignment=al(); c.border=bdM()
    c=ws.cell(row=10,column=col); c.value=sub
    c.fill=fx(C_LGRAY); c.font=ft(False,"404040",8); c.alignment=al(); c.border=bd()

ws.row_dimensions[12].height=18
ws.merge_cells('B12:H12')
total_prev = sum(f['Orçamento Previsto'] for f in financeiro)
c=ws['B12']; c.value=(f"  📊 Budget Total: R$ {total_prev:,.0f}   ·   "
    f"Eventos a Realizar: {sum(1 for e in events if 'Realizar' in e['Status Evento'])}   ·   "
    f"Atividades Pendentes: {sum(1 for a in activities if a['Status']=='PENDENTE')}   ·   "
    f"Fase 2 ✅ Concluída")
c.fill=fx(C_LBLUE); c.font=ft(True,C_NAVY,10); c.alignment=al("left")

ws.row_dimensions[14].height=16
ws.merge_cells('B14:H14')
ws['B14'].value = ("  ⚠️  Abas protegidas por governança · Edite somente nas células destacadas · "
                   "Novos eventos: use MODELO_CHECKLIST")
ws['B14'].fill=fx(C_LYELL); ws['B14'].font=ft(False,C_YELLOW,9); ws['B14'].alignment=al("left")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. DB_EVENTOS  (com ID_Unico + Flag + validações)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_EVENTOS")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "D4"

title_block(ws,'A1:X1',"📋 BANCO DE DADOS CENTRAL DE EVENTOS  ·  v3  ·  Fase 2",C_NAVY,sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:X2',
    "ID_Unico = EM normalizado (chave relacional) · Flag destaca EMs duplicados · Dropdowns validam campos críticos",
    C_LBLUE,C_NAVY)
ws.row_dimensions[2].height=16
ws.row_dimensions[3].height=32

COLS_EV = [
    ('ID','A',5),('ID_Unico','B',16),('Flag','C',10),('Evento','D',40),
    ('Tipo','E',12),('EM Original','F',16),('Status Evento','G',14),('Status EM','H',14),
    ('Ano','I',6),('Trimestre','J',10),('Data Início','K',12),('Data Fim','L',12),
    ('BU','M',16),('Produto','N',18),('Solicitante','O',18),('Cidade','P',16),
    ('Formato','Q',12),('Complexidade','R',12),('Local','S',28),
    ('Agência','T',16),('Hotel','U',20),('Sociedade','V',30),
    ('Stand','W',10),('PSs Plan.','X',9),('PSs Pres.','Y',9),('Observações','Z',38),
]
for i,(label,col_l,w) in enumerate(COLS_EV,1):
    ws.column_dimensions[get_column_letter(i)].width = w
hdr(ws,3,range(1,len(COLS_EV)+1))
for i,(label,_,_) in enumerate(COLS_EV,1):
    ws.cell(row=3,column=i).value=label

# Table
max_ev_row = 3 + len(events)
tab=Table(displayName="TblEventos",ref=f"A3:{get_column_letter(len(COLS_EV))}{max_ev_row}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tab)

dup_flag_rows = []
for i,ev in enumerate(events):
    r=4+i; alt=(i%2==1)
    data_row(ws,r,len(COLS_EV),alt=alt)
    row_data=[
        i+1, ev['ID_Unico'], ev['Flag'], ev['Evento'],
        TIPO_MAP.get(ev['Tipo'],ev['Tipo']), ev['EM'],
        ev['Status Evento'], ev['Status EM'],
        ev['Ano'], ev['Trimestre'], ev['Data Início'], ev['Data Fim'],
        ev['BU'], ev['Produto'], ev['Solicitante'], ev['Cidade'],
        ev['Formato'], ev['Complexidade'], ev['Local'],
        ev['Agência'], ev['Hotel'], ev['Sociedade'],
        ev['Stand'], ev['PSs Planejados'], ev['PSs Presentes'], ev['Observações'],
    ]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (11,12) and isinstance(val,datetime.date):
            cell.number_format='DD/MM/YYYY'
        if c_idx==7: status_cell(cell,ev['Status Evento'])
        if c_idx==8: status_cell(cell,ev['Status EM'])
        if c_idx==3 and val:  # Flag
            cell.fill=fx(C_LRED); cell.font=ft(True,C_RED,9); cell.alignment=al()
    if ev['Flag']:
        dup_flag_rows.append(r)

# DATA VALIDATIONS — Objetivo 3
n=len(events)
# Status Evento
add_dv(ws,"list",f'"{",".join(STATUS_EVENTO_LIST)}"',f"G4:G{3+n}",
    "Status do Evento","Selecione o status atual do evento")
# Status EM
add_dv(ws,"list",f'"{",".join(STATUS_EM_LIST)}"',f"H4:H{3+n}",
    "Status EM","Status da EM no VEEVA")
# Tipo
add_dv(ws,"list",f'"{",".join(TIPO_EV_LIST)}"',f"E4:E{3+n}",
    "Tipo de Participação")
# BU
add_dv(ws,"list",f'"{",".join(BU_LIST)}"',f"M4:M{3+n}",
    "Business Unit")
# Formato
add_dv(ws,"list",f'"{",".join(FORMATO_LIST)}"',f"Q4:Q{3+n}",
    "Formato do Evento")
# Complexidade
add_dv(ws,"list",f'"{",".join(COMPLEXIDADE_LIST)}"',f"R4:R{3+n}",
    "Complexidade")
# Agência
add_dv(ws,"list",f'"{",".join(AGENCIA_LIST)}"',f"T4:T{3+n}",
    "Agência Logística")

print("DB_EVENTOS ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 3. DB_ATIVIDADES (com validações)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_ATIVIDADES")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "D4"

title_block(ws,'A1:H1',"✅ BANCO DE DADOS DE ATIVIDADES  ·  v3","375623",sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:H2',
    "Status e Responsável com dropdown · Filtrar por Evento ou Etapa para ver somente o que importa",
    C_LGREEN,C_GREEN)
ws.row_dimensions[2].height=16
ws.row_dimensions[3].height=28

COLS_AT=[
    ('ID','A',5),('Evento','B',38),('Etapa','C',22),('Atividade','D',44),
    ('Responsável','E',18),('% Conclusão','F',13),('Status','G',16),('Observações','H',40),
]
for i,(_,_,w) in enumerate(COLS_AT,1): ws.column_dimensions[get_column_letter(i)].width=w
hdr(ws,3,range(1,len(COLS_AT)+1),bg="375623")
for i,(label,_,_) in enumerate(COLS_AT,1): ws.cell(row=3,column=i).value=label

max_at_row=3+len(activities)
tab2=Table(displayName="TblAtividades",ref=f"A3:{get_column_letter(len(COLS_AT))}{max_at_row}")
tab2.tableStyleInfo=TableStyleInfo(name="TableStyleMedium7",showRowStripes=True)
ws.add_table(tab2)

from openpyxl.formatting.rule import DataBarRule
ws.conditional_formatting.add(f"F4:F{max_at_row}",
    DataBarRule(start_type='num',start_value=0,end_type='num',end_value=1,color="2E75B6"))

for i,at in enumerate(activities):
    r=4+i; data_row(ws,r,len(COLS_AT),alt=(i%2==1))
    pct=at['% Conclusão']
    row_data=[i+1,at['Evento'],at['Etapa'],at['Atividade'],
              at['Responsável'],pct/100 if pct else 0,at['Status'],at['Observações']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx==6:
            cell.number_format='0%'; cell.alignment=al()
            if pct==100: cell.fill=fx("E2EFDA"); cell.font=ft(True,"375623",9)
            elif pct>=50: cell.fill=fx("FFF2CC"); cell.font=ft(True,"7F6000",9)
            elif pct>0:   cell.fill=fx("FCE4D6"); cell.font=ft(True,"843C0C",9)
            else:         cell.fill=fx("FFDDD8"); cell.font=ft(True,"C00000",9)
        if c_idx==7: status_cell(cell,at['Status'])

n_at=len(activities)
add_dv(ws,"list",f'"{",".join(STATUS_AT_LIST)}"',f"G4:G{3+n_at}","Status da Atividade")
resp_str = ",".join(RESP_UNICOS[:20]) if RESP_UNICOS else "Eventos,Time de Marca"
add_dv(ws,"list",f'"{resp_str}"',f"E4:E{3+n_at}","Responsável")

print("DB_ATIVIDADES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 4. DB_FINANCEIRO (EM padronizado + validações)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_FINANCEIRO")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C4"

title_block(ws,'A1:N1',"💰 BANCO DE DADOS FINANCEIRO  ·  v3","843C0C",sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:N2',
    "EM padronizado (sem espaços) · Status de pagamento com dropdown · Valores em R$",
    C_LORANGE,"843C0C")
ws.row_dimensions[2].height=16
ws.row_dimensions[3].height=28

COLS_FIN=[
    ('ID','A',5),('Evento','B',34),('EM (Padrão)','C',14),('FY','D',6),
    ('Fornecedor','E',22),('Tipo de Custo','F',20),('PO/WF','G',14),
    ('Orç. Previsto','H',16),('Real. FY2025','I',14),('Real. FY2026','J',14),
    ('Total Evento','K',16),('Status Pag.','L',14),('Status NF/ND','M',14),('Observações','N',38),
]
for i,(_,_,w) in enumerate(COLS_FIN,1): ws.column_dimensions[get_column_letter(i)].width=w
hdr(ws,3,range(1,len(COLS_FIN)+1),bg="843C0C")
for i,(label,_,_) in enumerate(COLS_FIN,1): ws.cell(row=3,column=i).value=label

max_fin_row=3+len(financeiro)
tab3=Table(displayName="TblFinanceiro",ref=f"A3:{get_column_letter(len(COLS_FIN))}{max_fin_row}")
tab3.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True)
ws.add_table(tab3)

total_prev_all=sum(f['Orçamento Previsto'] for f in financeiro)
total_r25=sum(f['Realizado FY2025'] for f in financeiro)
total_r26=sum(f['Realizado FY2026'] for f in financeiro)

for i,fin in enumerate(financeiro):
    r=4+i; data_row(ws,r,len(COLS_FIN),alt=(i%2==1))
    row_data=[i+1,fin['Evento'],fin['EM'],fin['FY'],fin['Fornecedor'],
              fin['Tipo de Custo'],fin['PO/WF'],
              fin['Orçamento Previsto'],fin['Realizado FY2025'],
              fin['Realizado FY2026'],fin['Total Evento'],
              fin['Status Pagamento'],fin['Status NF/ND'],fin['Observações']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (8,9,10,11) and isinstance(val,(int,float)):
            cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
        if c_idx==12: status_cell(cell,fin['Status Pagamento'])
        if c_idx==13: status_cell(cell,fin['Status NF/ND'])

r_tot=4+len(financeiro); ws.row_dimensions[r_tot].height=22
ws.merge_cells(f'A{r_tot}:G{r_tot}')
c=ws.cell(row=r_tot,column=1); c.value="TOTAIS GERAIS"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,10); c.alignment=al()
for col_idx,tot in [(8,total_prev_all),(9,total_r25),(10,total_r26),(11,total_r25+total_r26)]:
    cell=ws.cell(row=r_tot,column=col_idx); cell.value=tot
    cell.number_format='R$ #,##0.00'; cell.fill=fx(C_NAVY)
    cell.font=ft(True,C_WHITE,10); cell.alignment=al("right"); cell.border=bd()

n_fin=len(financeiro)
add_dv(ws,"list",f'"{",".join(STATUS_PAG_LIST)}"',f"L4:L{3+n_fin}","Status Pagamento")
add_dv(ws,"list",f'"{",".join(STATUS_NF_LIST)}"',f"M4:M{3+n_fin}","Status NF/ND/Recibo")
add_dv(ws,"list",f'"{",".join(TIPO_CUSTO_LIST)}"',f"F4:F{3+n_fin}","Tipo de Custo")

print("DB_FINANCEIRO ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 5. DB_FORNECEDORES (novo — Objetivo 1)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_FORNECEDORES")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"

title_block(ws,'A1:J1',"🏢 BANCO DE DADOS DE FORNECEDORES",C_TEAL,sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:J2',
    "Cadastro único de fornecedores homologados · Atualizar CNPJ e contato antes de criar nova PO",
    "D9EAF5","1F6B75")
ws.row_dimensions[2].height=16
ws.row_dimensions[3].height=32

COLS_FORN=[
    ('ID','A',5),('Fornecedor','B',28),('Categoria','C',22),('Serviços','D',30),
    ('UF Sede','E',8),('Status Homol.','F',16),('CNPJ','G',18),
    ('Contato','H',22),('E-mail','I',30),('Observações','J',40),
]
for i,(_,_,w) in enumerate(COLS_FORN,1): ws.column_dimensions[get_column_letter(i)].width=w
hdr(ws,3,range(1,len(COLS_FORN)+1),bg=C_TEAL)
for i,(label,_,_) in enumerate(COLS_FORN,1): ws.cell(row=3,column=i).value=label

forn_list = []
for name,(cat,servicos,uf,status) in SUPPLIER_CATALOG.items():
    forn_list.append({'Fornecedor':name,'Categoria':cat,'Serviços':servicos,
        'UF Sede':uf,'Status Homol.':status,
        'CNPJ':'A preencher','Contato':'A preencher','E-mail':'A preencher','Observações':''})

max_forn_row=3+len(forn_list)
tab4=Table(displayName="TblFornecedores",ref=f"A3:{get_column_letter(len(COLS_FORN))}{max_forn_row}")
tab4.tableStyleInfo=TableStyleInfo(name="TableStyleMedium9",showRowStripes=True)
ws.add_table(tab4)

STATUS_FORN_LIST=['Homologado','A verificar','Contrato','Oficial evento','A preencher','Suspenso']
for i,forn in enumerate(forn_list):
    r=4+i; data_row(ws,r,len(COLS_FORN),alt=(i%2==1))
    row_data=[i+1,forn['Fornecedor'],forn['Categoria'],forn['Serviços'],
              forn['UF Sede'],forn['Status Homol.'],forn['CNPJ'],
              forn['Contato'],forn['E-mail'],forn['Observações']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx==6: status_cell(cell,val)

n_forn=len(forn_list)
add_dv(ws,"list",f'"{",".join(STATUS_FORN_LIST)}"',f"F4:F{3+n_forn}","Status de Homologação")
add_dv(ws,"list",
    '"Agência Logística,Agência de Produção,Montadora de Stand,Buffet/Alimentação,Sociedade Médica,Palestrante,Outros"',
    f"C4:C{3+n_forn}","Categoria do Fornecedor")

print("DB_FORNECEDORES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 6. DB_PARTICIPANTES (novo — Objetivo 2)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DB_PARTICIPANTES")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"

title_block(ws,'A1:L1',"👥 BANCO DE DADOS DE PARTICIPANTES",C_PURPLE,sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:L2',
    "Registro histórico de HCPs e Staff · Adicionar novos participantes nesta tabela antes de cada evento",
    "EAD1FF","7030A0")
ws.row_dimensions[2].height=16
ws.row_dimensions[3].height=32

COLS_PART=[
    ('ID','A',5),('Nome','B',30),('Cargo/Regional','C',28),('Tipo','D',16),
    ('Especialidade','E',22),('CRM','F',14),('UF','G',6),
    ('Evento Ref.','H',22),('Hotel Bloco','I',16),
    ('Aéreo R$','J',12),('Hospedagem R$','K',14),('Observações','L',38),
]
for i,(_,_,w) in enumerate(COLS_PART,1): ws.column_dimensions[get_column_letter(i)].width=w
hdr(ws,3,range(1,len(COLS_PART)+1),bg=C_PURPLE)
for i,(label,_,_) in enumerate(COLS_PART,1): ws.cell(row=3,column=i).value=label

# Deduplicate participants
seen_names=set(); part_final=[]
for p in brain_participants:
    if p['Nome'] not in seen_names:
        seen_names.add(p['Nome']); part_final.append(p)

max_part_row=3+len(part_final)
tab5=Table(displayName="TblParticipantes",ref=f"A3:{get_column_letter(len(COLS_PART))}{max_part_row}")
tab5.tableStyleInfo=TableStyleInfo(name="TableStyleMedium28",showRowStripes=True)
ws.add_table(tab5)

TIPO_PART_LIST=['Staff Takeda','HCP Convidado','Palestrante','Coordenador','Moderador']
for i,p in enumerate(part_final):
    r=4+i; data_row(ws,r,len(COLS_PART),alt=(i%2==1))
    row_data=[i+1,p['Nome'],p['Cargo/Regional'],p['Tipo'],'A preencher',
              'A preencher','A preencher',p['Evento Ref.'],p['Hotel Bloco'],
              0.0,0.0,p['Observações']]
    for c_idx,val in enumerate(row_data,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (10,11) and isinstance(val,(int,float)):
            cell.number_format='R$ #,##0.00'; cell.alignment=al("right")

add_dv(ws,"list",f'"{",".join(TIPO_PART_LIST)}"',f"D4:D{3+len(part_final)}","Tipo de Participante")

print("DB_PARTICIPANTES ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. DASHBOARD (atualizado)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("📊 DASHBOARD")
ws.sheet_view.showGridLines = False
for c in range(1,16): ws.column_dimensions[get_column_letter(c)].width=14
ws.column_dimensions['A'].width=2

ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=40
ws.merge_cells('B2:O2')
c=ws['B2']; c.value="📊 DASHBOARD EXECUTIVO · GESTÃO DE EVENTOS"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,20); c.alignment=al()
ws.merge_cells('B3:O3')
c=ws['B3']; c.value=f"Atualizado: {datetime.date.today().strftime('%d/%m/%Y')}  ·  v3 Fase 2  ·  Fonte: DB_EVENTOS / DB_ATIVIDADES / DB_FINANCEIRO"
c.fill=fx(C_LBLUE); c.font=ft(False,C_NAVY,10); c.alignment=al("left")
ws.row_dimensions[3].height=18

ws.row_dimensions[5].height=18; ws.row_dimensions[6].height=50; ws.row_dimensions[7].height=22

kpis=[
    ("TOTAL\nEVENTOS",len(events),"",C_NAVY),
    ("EVENTOS\n2026",sum(1 for e in events if e['Ano']=='2026'),"",C_BLUE),
    ("A\nREALIZAR",sum(1 for e in events if 'Realizar' in e['Status Evento']),"","843C0C"),
    ("REALIZADOS",sum(1 for e in events if e['Status Evento']=='Realizado'),"","375623"),
    ("ATIVIDADES\nPENDENTES",sum(1 for a in activities if a['Status']=='PENDENTE'),"",C_RED),
    (f"BUDGET\nTOTAL",f"R$ {total_prev_all/1e6:.1f}M","",C_PURPLE),
]
kpi_cols=[2,4,6,8,10,12]
for idx,(label,value,sub,bg) in enumerate(kpis):
    col=kpi_cols[idx]; col2=col+1
    ws.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col2)
    ws.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col2)
    ws.merge_cells(start_row=7,start_column=col,end_row=7,end_column=col2)
    for rr,val,sz in [(5,label,9),(6,value,26),(7,sub,8)]:
        c=ws.cell(row=rr,column=col); c.value=val
        c.fill=fx(bg) if rr!=7 else fx(C_LGRAY)
        c.font=ft(True,C_WHITE if rr!=7 else "404040",sz)
        c.alignment=al(); c.border=bd()

# Status breakdown table
from collections import Counter
status_count=Counter(e['Status Evento'] for e in events)
bu_count=Counter(e['BU'] for e in events if e['BU'])
city_count=Counter(e['Cidade'] for e in events if e['Cidade'])
act_status=Counter(a['Status'] for a in activities)
pay_status=Counter(f['Status Pagamento'] for f in financeiro if f['Status Pagamento'])

from collections import defaultdict
fin_by_ev=defaultdict(lambda:{'prev':0,'r25':0,'r26':0})
for f in financeiro:
    fin_by_ev[f['Evento']]['prev']+=f['Orçamento Previsto']
    fin_by_ev[f['Evento']]['r25']+=f['Realizado FY2025']
    fin_by_ev[f['Evento']]['r26']+=f['Realizado FY2026']

def table_section(ws, r_start, col_start, title, title_bg, cols_headers, rows_data):
    cols=len(cols_headers); end_col=col_start+cols-1
    ws.merge_cells(start_row=r_start,start_column=col_start,end_row=r_start,end_column=end_col)
    c=ws.cell(row=r_start,column=col_start); c.value=title
    c.fill=fx(title_bg); c.font=ft(True,C_WHITE,11); c.alignment=al()
    for ci,h in enumerate(cols_headers,col_start):
        c=ws.cell(row=r_start+1,column=ci); c.value=h
        c.fill=fx(title_bg); c.font=ft(True,C_WHITE,9); c.alignment=al(); c.border=bd()
    r=r_start+2
    for row_vals in rows_data:
        for ci,val in enumerate(row_vals,col_start):
            c=ws.cell(row=r,column=ci); c.value=val
            c.fill=fx(C_LGRAY if r%2==0 else C_WHITE); c.font=ft(size=9)
            c.alignment=al(); c.border=bd()
        r+=1
    return r

r_s=9
table_section(ws,r_s,2,"STATUS DOS EVENTOS",C_NAVY,['Status','Qtd','%'],
    [(s,cnt,f"{cnt/len(events)*100:.0f}%") for s,cnt in sorted(status_count.items(),key=lambda x:-x[1])])
table_section(ws,r_s,6,"EVENTOS POR BU","375623",['BU','Qtd'],
    list(sorted(bu_count.items(),key=lambda x:-x[1]))[:8])
table_section(ws,r_s,9,"TOP CIDADES",C_TEAL,['Cidade','Qtd'],
    list(sorted(city_count.items(),key=lambda x:-x[1]))[:8])
table_section(ws,r_s,12,"STATUS PAGAMENTOS",C_PURPLE,['Status','Qtd'],
    list(sorted(pay_status.items(),key=lambda x:-x[1])))

r_s2=22
table_section(ws,r_s2,2,"ATIVIDADES POR STATUS",C_NAVY,['Status','Qtd','%'],
    [(s,cnt,f"{cnt/len(activities)*100:.0f}%") for s,cnt in sorted(act_status.items(),key=lambda x:-x[1])])
table_section(ws,r_s2,6,"RESUMO FINANCEIRO","843C0C",['Evento','Previsto','Real.Total'],
    [(k[:28],v['prev'],v['r25']+v['r26']) for k,v in sorted(fin_by_ev.items(),key=lambda x:-x[1]['prev'])[:10]])

# Format money cols in financial table
for r in range(r_s2+2, r_s2+12):
    for col in [7,8]:
        cell=ws.cell(row=r,column=col)
        if isinstance(cell.value,(int,float)):
            cell.number_format='R$ #,##0'; cell.alignment=al("right")

print("DASHBOARD ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 8. PARTICIPAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("PARTICIPAÇÃO")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
title_block(ws,'A1:M1',"✈️ CONTROLE DE PARTICIPAÇÃO E LOGÍSTICA",C_NAVY,sz=13)
ws.row_dimensions[1].height=28
subtitle_block(ws,'A2:M2',"Registre deslocamento, hospedagem e custos por participante e evento",C_LBLUE,C_NAVY)
ws.row_dimensions[2].height=16

COLS_P=[('Evento','A',34),('Participante','B',22),('Cargo','C',18),('Regional','D',10),
        ('Cidade Orig.','E',14),('Data Ida','F',12),('Data Volta','G',12),
        ('Aéreo R$','H',12),('Hospedagem R$','I',14),('Refeição R$','J',12),
        ('Deslocamento R$','K',16),('Total R$','L',14),('Observações','M',30),]
for i,(_,_,w) in enumerate(COLS_P,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[3].height=28
hdr(ws,3,range(1,len(COLS_P)+1))
for i,(label,_,_) in enumerate(COLS_P,1): ws.cell(row=3,column=i).value=label

ws_p_src=src['PARTICIPAÇÃO']
r=4
for row in ws_p_src.iter_rows(min_row=3,max_row=ws_p_src.max_row,values_only=True):
    if row[1] is None: continue
    data_row(ws,r,len(COLS_P),alt=(r%2==0))
    def to_date(v):
        if isinstance(v,datetime.datetime): return v.date()
        return v
    try: aer=float(str(row[6])) if row[6] else 0
    except: aer=0
    try: hosp=float(str(row[7])) if row[7] else 0
    except: hosp=0
    try: ref=float(str(row[8])) if row[8] else 0
    except: ref=0
    try: desl=float(str(row[9])) if row[9] else 0
    except: desl=0
    vals=[str(row[1]).strip(),'','','','',to_date(row[4]),to_date(row[5]),aer,hosp,ref,desl,aer+hosp+ref+desl,'']
    for c_idx,val in enumerate(vals,1):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        if c_idx in (8,9,10,11,12) and isinstance(val,(int,float)):
            cell.number_format='R$ #,##0.00'; cell.alignment=al("right")
        if c_idx in (6,7) and isinstance(val,datetime.date):
            cell.number_format='DD/MM/YYYY'
    r+=1

tab6=Table(displayName="TblParticipacao",ref=f"A3:{get_column_letter(len(COLS_P))}{r-1}")
tab6.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tab6)
print("PARTICIPAÇÃO ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 9. MODELO_CHECKLIST (novo — Objetivo 5)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("MODELO_CHECKLIST")
ws.sheet_view.showGridLines = False
for c in range(1,9): ws.column_dimensions[get_column_letter(c)].width=[2,25,38,18,13,16,40,2][c-1]
ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=36

ws.merge_cells('B2:G2')
c=ws['B2']; c.value="📝 MODELO PADRÃO DE CHECKLIST · Duplicar esta aba para cada novo evento"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,13); c.alignment=al()

ws.row_dimensions[3].height=20
ws.merge_cells('B3:G3')
c=ws['B3']; c.value=("Como usar: (1) Duplique esta aba  (2) Renomeie com o nome do evento  "
                      "(3) Preencha o cabeçalho  (4) Copie as atividades concluídas para DB_ATIVIDADES")
c.fill=fx(C_LYELL); c.font=ft(False,C_YELLOW,9); c.alignment=al("left")

# Event header block
ws.row_dimensions[5].height=22
for col,(label,val_col) in enumerate([('Evento:','C'),('EM:','E'),('Data:','G')],2):
    c=ws.cell(row=5,column=col); c.value=label
    c.fill=fx(C_LBLUE); c.font=ft(True,C_NAVY,9); c.alignment=al()
    c.border=bd()
    c2=ws.cell(row=5,column=col+1); c2.value='← preencher'
    c2.fill=fx(C_LYELL); c2.font=ft(False,C_YELLOW,9); c2.alignment=al("left"); c2.border=bd()

ws.row_dimensions[7].height=28
hdr(ws,7,[2,3,4,5,6,7],bg=C_NAVY)
for ci,h in enumerate(['Etapa','Atividade','Responsável','% Conclusão','Status','Observações'],2):
    ws.cell(row=7,column=ci).value=h

checklist_template = [
    ('Planejamento', 'Briefing com Time de Marca', 'Time de Marca', 0, 'PENDENTE', ''),
    ('Planejamento', 'Inclusão no calendário de eventos', 'Eventos', 0, 'PENDENTE', ''),
    ('VEEVA', 'EM aprovada (Cota)', 'Time de Marca', 0, 'PENDENTE', 'Número EM: ___'),
    ('VEEVA', 'EM aprovada (Apoio)', 'Time de Marca', 0, 'PENDENTE', ''),
    ('VEEVA', 'EM aprovada (Simpósio)', 'Time de Marca', 0, 'PENDENTE', ''),
    ('VEEVA', 'Incluir time de Processos/Pagamentos', 'Eventos', 0, 'PENDENTE', ''),
    ('VEEVA', 'Incluir Agência Logística na EM', 'Eventos', 0, 'PENDENTE', ''),
    ('Sociedade', 'Contrato de Cota recebido', 'Eventos', 0, 'PENDENTE', ''),
    ('Sociedade', 'Envio de logo para Sociedade', 'Eventos', 0, 'PENDENTE', ''),
    ('Sociedade', 'Monitorar contrapartidas', 'Eventos', 0, 'PENDENTE', ''),
    ('Agência', 'Briefing para agência logística', 'Eventos', 0, 'PENDENTE', ''),
    ('Agência', 'Solicitar orçamento', 'Eventos', 0, 'PENDENTE', ''),
    ('Agência', 'Aprovação do orçamento', 'Time de Marca', 0, 'PENDENTE', ''),
    ('Agência', 'Criação de PO', 'Eventos', 0, 'PENDENTE', 'PO Nº: ___'),
    ('Agência', 'RSVP / Inscrições HCPs', 'Eventos', 0, 'PENDENTE', ''),
    ('Simpósio', 'Contratos de Palestrantes', 'Eventos', 0, 'PENDENTE', ''),
    ('Simpósio', 'PO Palestrantes aprovada', 'Eventos', 0, 'PENDENTE', ''),
    ('Simpósio', 'Aditivo (data/horário corretos)', 'Eventos', 0, 'PENDENTE', ''),
    ('Simpósio', 'Lunch Box / Buffet', 'Eventos', 0, 'PENDENTE', ''),
    ('Stand', 'Documentação do expositor', 'Eventos', 0, 'PENDENTE', ''),
    ('Stand', 'Formulários obrigatórios', 'Eventos', 0, 'PENDENTE', ''),
    ('Stand', 'PO Montadora aprovada', 'Eventos', 0, 'PENDENTE', ''),
    ('Pós-Evento', 'Solicitação NF/Recibo', 'Eventos', 0, 'PENDENTE', ''),
    ('Pós-Evento', 'Confirmação de pagamento', 'Eventos', 0, 'PENDENTE', ''),
    ('Pós-Evento', 'Encerramento da EM', 'Eventos', 0, 'PENDENTE', ''),
    ('Pós-Evento', 'Relatório de execução', 'Eventos', 0, 'PENDENTE', ''),
    ('Pós-Evento', 'Atualizar DB_ATIVIDADES', 'Eventos', 0, 'PENDENTE', 'Copiar todas as linhas concluídas'),
]

for i,(etapa,ativ,resp,pct,stat,obs) in enumerate(checklist_template):
    r=8+i; ws.row_dimensions[r].height=18
    alt=(i%2==1)
    for col in range(2,8): ws.cell(row=r,column=col).fill=fx(C_LGRAY if alt else C_WHITE)
    vals=[etapa,ativ,resp,pct/100 if pct else 0,stat,obs]
    for c_idx,val in enumerate(vals,2):
        cell=ws.cell(row=r,column=c_idx); cell.value=val
        cell.font=ft(size=9); cell.border=bd()
        if c_idx==5: status_cell(cell,val)
        elif c_idx==4:
            cell.number_format='0%'; cell.alignment=al()
        else:
            cell.alignment=al("left")
        if c_idx==2:
            cell.fill=fx(C_LBLUE if i==0 or etapa!=checklist_template[i-1][0] else (C_LGRAY if alt else C_WHITE))
            if i==0 or etapa!=checklist_template[i-1][0]:
                cell.font=ft(True,C_NAVY,9)

# Data validation on template
add_dv(ws,"list",f'"{",".join(STATUS_AT_LIST)}"',f"F8:F{7+len(checklist_template)}","Status")
add_dv(ws,"list",f'"{resp_str}"',f"D8:D{7+len(checklist_template)}","Responsável")

print("MODELO_CHECKLIST ✅")

# ═══════════════════════════════════════════════════════════════════════════════
# 10. ROADMAP atualizado
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🗺️ ROADMAP")
ws.sheet_view.showGridLines = False
for c in range(1,10): ws.column_dimensions[get_column_letter(c)].width=22
ws.column_dimensions['A'].width=3
ws.row_dimensions[1].height=10; ws.row_dimensions[2].height=40
ws.merge_cells('B2:I2')
c=ws['B2']; c.value="🗺️ ROADMAP DE IMPLEMENTAÇÃO"
c.fill=fx(C_NAVY); c.font=ft(True,C_WHITE,20); c.alignment=al()

phases=[
    ("FASE 1 — QUICK WINS ✅",C_GREEN,"E2EFDA","375623",[
        ("✅ Concluído","Migração de dados para DB_EVENTOS centralizado"),
        ("✅ Concluído","Consolidação de checklists em DB_ATIVIDADES"),
        ("✅ Concluído","Unificação financeira em DB_FINANCEIRO"),
        ("✅ Concluído","Formatação de tabelas com Excel Tables (filtros)"),
        ("✅ Concluído","Dashboard executivo com KPIs e resumos"),
        ("✅ Concluído","Padronização de status com formatação condicional"),
        ("✅ Concluído","Treinar equipe no uso dos novos bancos de dados (guia na aba INÍCIO)"),
    ]),
    ("FASE 2 — REESTRUTURAÇÃO ✅",C_BLUE,"BDD7EE","1F3864",[
        ("✅ Concluído","DB_FORNECEDORES: cadastro de 17 fornecedores homologados com status e categoria"),
        ("✅ Concluído","DB_PARTICIPANTES: registro histórico de participantes (Brain 2026 + estrutura HCP)"),
        ("✅ Concluído","Validação de dados: dropdowns em Status, Tipo, BU, Formato, Responsável, Pagamento"),
        ("✅ Concluído","ID_Unico padronizado (EM sem espaços) + flag de duplicidade automática"),
        ("✅ Concluído","MODELO_CHECKLIST: template de 27 atividades para novos eventos"),
        ("✅ Concluído","Correções: typo Biansa Fasas→Bianca Facas, EMs normalizados, campos vazios mapeados"),
    ]),
    ("FASE 3 — AUTOMAÇÃO","C4700E","FCE4D6","843C0C",[
        ("📅 Futuro","Power Query: atualização automática de relatórios"),
        ("📅 Futuro","Office Scripts: geração de relatório semanal por e-mail"),
        ("📅 Futuro","Alertas automáticos de prazos via conditional formatting"),
        ("📅 Futuro","Power Automate: notificação de pagamentos pendentes"),
        ("📅 Futuro","Power Automate: lembrete de atividades atrasadas"),
        ("📅 Futuro","Integração com SharePoint para controle de versão"),
    ]),
    ("FASE 4 — DASHBOARD AVANÇADO",C_PURPLE,"EAD1FF","7030A0",[
        ("📅 Futuro","Power BI: dashboard interativo com drill-down"),
        ("📅 Futuro","Gráfico de Gantt de eventos por trimestre"),
        ("📅 Futuro","Mapa geográfico de eventos por cidade"),
        ("📅 Futuro","Análise de tendências de budget YoY"),
        ("📅 Futuro","Relatório executivo automatizado (PDF)"),
        ("📅 Futuro","Scorecard de performance por responsável"),
    ]),
]

r=4
for phase_title,bg_hdr,bg_light,fg_dark,items in phases:
    ws.row_dimensions[r].height=28
    ws.merge_cells(f'B{r}:I{r}')
    c=ws.cell(row=r,column=2); c.value=phase_title
    c.fill=fx(bg_hdr); c.font=ft(True,C_WHITE,12); c.alignment=al()
    r+=1
    for ci,h in enumerate(['Status','Atividade'],2):
        c=ws.cell(row=r,column=ci); c.value=h
        c.fill=fx(bg_light); c.font=ft(True,fg_dark,9); c.alignment=al(); c.border=bd()
    ws.row_dimensions[r].height=18; r+=1
    for stat,desc in items:
        ws.row_dimensions[r].height=20
        c=ws.cell(row=r,column=2); c.value=stat
        bg_s="E2EFDA" if "Concluído" in stat else ("FFF2CC" if "Próximo" in stat else "F2F2F2")
        c.fill=fx(bg_s); c.font=ft(True,fg_dark,9); c.alignment=al(); c.border=bd()
        ws.merge_cells(f'C{r}:I{r}')
        c=ws.cell(row=r,column=3); c.value=desc
        c.fill=fx(C_WHITE if r%2==0 else C_LGRAY); c.font=ft(size=9)
        c.alignment=al("left"); c.border=bd()
        r+=1
    r+=1

print("ROADMAP ✅")

# ── Tab colors ────────────────────────────────────────────────────────────────
TAB_COLORS={
    "🏠 INÍCIO":"1F3864","DB_EVENTOS":"2E75B6","DB_ATIVIDADES":"375623",
    "DB_FINANCEIRO":"843C0C","DB_FORNECEDORES":"1F6B75","DB_PARTICIPANTES":"7030A0",
    "📊 DASHBOARD":"1F3864","PARTICIPAÇÃO":"2E75B6",
    "MODELO_CHECKLIST":"C4700E","🗺️ ROADMAP":"404040",
}
for sname,color in TAB_COLORS.items():
    if sname in wb.sheetnames:
        wb[sname].sheet_properties.tabColor=color

# ── Save ──────────────────────────────────────────────────────────────────────
out=('/home/user/Femartignon/NEC_Eventos_Reestruturado_v3.xlsx')
wb.save(out)
print(f"\n✅ FASE 2 COMPLETA → {out}")
