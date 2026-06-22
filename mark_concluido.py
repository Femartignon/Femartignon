"""Marca como CONCLUÍDO todas as atividades vencidas (prazo < hoje) que estavam
PENDENTE/EM ANDAMENTO, e ajusta % Conclusão para 100%."""
import datetime
from openpyxl import load_workbook

PATH = 'NEC_Eventos_Reestruturado_v4.xlsx'
HOJE = datetime.date(2026, 6, 22)
wb = load_workbook(PATH)
at = wb['DB_ATIVIDADES']

alterados = 0
for r in range(4, at.max_row+1):
    st = at.cell(row=r, column=8).value
    f = at.cell(row=r, column=6).value
    if not f or str(st).strip() not in ('PENDENTE', 'EM ANDAMENTO'):
        continue
    d = f.date() if isinstance(f, datetime.datetime) else f
    if d < HOJE:
        at.cell(row=r, column=8).value = 'CONCLUÍDO'   # Status
        g = at.cell(row=r, column=7)                    # % Conclusão -> 100%
        g.value = 1
        g.number_format = '0%'
        alterados += 1

wb.save(PATH)
print(f"✅ {alterados} atividades vencidas marcadas como CONCLUÍDO (100%)")

# Verificação
wb2 = load_workbook(PATH)
at2 = wb2['DB_ATIVIDADES']
from collections import Counter
status = Counter()
venc_restante = 0
for r in range(4, at2.max_row+1):
    s = at2.cell(row=r, column=8).value
    f = at2.cell(row=r, column=6).value
    if at2.cell(row=r, column=2).value is None: continue
    status[str(s).strip() if s else '(vazio)'] += 1
    if f and str(s).strip() in ('PENDENTE','EM ANDAMENTO'):
        d = f.date() if isinstance(f, datetime.datetime) else f
        if d < HOJE: venc_restante += 1
print("Status agora:", dict(status))
print("Vencidos ainda em aberto:", venc_restante)
