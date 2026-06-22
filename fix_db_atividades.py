"""
Fix in-place: DB_ATIVIDADES no arquivo enviado pelo usuário.
  - Coluna F (Prazo): formatar como DATA (DD/MM/AAAA), limpar texto "A definir",
    adicionar validação de data (bloqueia texto, permite vazio)
  - Coluna G (% Conclusão): reforçar formato 0% em toda a coluna
  - Formatação condicional: LIMPAR tudo (estava fragmentada/órfã) e reconstruir limpo
    sobre o range completo A4:I<fim>, com DataBar por último
NÃO regenera o arquivo do zero — preserva os dados do usuário.
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidation

PATH = 'NEC_Eventos_Reestruturado_v4.xlsx'
wb = load_workbook(PATH)
ws = wb['DB_ATIVIDADES']

# Cores (mesma paleta da v4)
C_LRED, C_RED       = "FFDDD8", "C00000"
C_LORANGE, C_ORANGE = "FCE4D6", "843C0C"
C_LYELL, C_YELLOW   = "FFF2CC", "7F6000"
C_LBLUE, C_NAVY     = "BDD7EE", "1F3864"

def fx(h): return PatternFill("solid", fgColor=h)
def ft(bold, color): return Font(bold=bold, color=color, size=9, name="Calibri")

# Descobrir extensão real dos dados (a tabela vai até a última linha)
tbl = ws.tables['TblAtividades']
last_row = int(tbl.ref.split(':')[1][1:])   # ex 'A3:I146' -> 146
first = 4
print(f"Tabela TblAtividades: {tbl.ref} | data rows {first}..{last_row}")

# ── 1. Coluna F (Prazo): limpar "A definir", formatar como data ───────────────
limpos = 0
for r in range(first, last_row+1):
    cell = ws.cell(row=r, column=6)
    if isinstance(cell.value, str) and cell.value.strip().lower() in ('a definir', 'a definr', ''):
        cell.value = None
        limpos += 1
    cell.number_format = 'DD/MM/YYYY'
print(f"Coluna F: {limpos} células 'A definir' limpas e formatadas como data")

# ── 2. Coluna G (% Conclusão): reforçar 0% ────────────────────────────────────
for r in range(first, last_row+1):
    ws.cell(row=r, column=7).number_format = '0%'
print("Coluna G: formato 0% reaplicado")

# ── 3. Limpar TODA a formatação condicional bagunçada ─────────────────────────
ws.conditional_formatting = ConditionalFormattingList()
print("Formatação condicional antiga removida")

# ── 4. Reconstruir CF limpo sobre o range completo ────────────────────────────
def add_cf(rng, formula, bg, fg):
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[formula], fill=fx(bg), font=ft(True, fg), stopIfTrue=False))

cf_rng = f"A{first}:I{last_row}"
# Status H, Prazo F. Anchor na primeira linha de dados (4).
add_cf(cf_rng, f'=$H{first}="ATRASADO"', C_LRED, C_RED)
add_cf(cf_rng, f'=AND($H{first}="PENDENTE",ISNUMBER($F{first}),$F{first}<TODAY())', C_LRED, C_RED)
add_cf(cf_rng, f'=AND($H{first}="PENDENTE",ISNUMBER($F{first}),$F{first}>=TODAY(),$F{first}<=TODAY()+7)', C_LORANGE, C_ORANGE)
add_cf(cf_rng, f'=$H{first}="PENDENTE"', C_LYELL, C_YELLOW)
add_cf(cf_rng, f'=$H{first}="EM ANDAMENTO"', C_LBLUE, C_NAVY)
add_cf(cf_rng, f'=$H{first}="CONCLUÍDO"', "F0FFF0", "375623")
# DataBar % por último, só na coluna G
ws.conditional_formatting.add(f"G{first}:G{last_row}",
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color="2E75B6"))
print(f"CF reconstruído limpo em {cf_rng} + DataBar em G")

# ── 5. Validação de data na coluna Prazo (bloqueia texto, permite vazio) ───────
# Remover validações antigas que toquem F
keep = [dv for dv in ws.data_validations.dataValidation if 'F' not in str(dv.sqref).split(':')[0][0:1] or True]
# (mantemos as existentes; apenas adicionamos a de data em F)
dv_date = DataValidation(
    type="date", operator="greaterThanOrEqual", formula1="43831",  # 01/01/2020
    allow_blank=True, showInputMessage=True, showErrorMessage=True)
dv_date.promptTitle = "Prazo da Atividade"
dv_date.prompt = "Insira a data no formato DD/MM/AAAA.\nDeixe em branco se ainda não houver prazo definido."
dv_date.errorTitle = "Data inválida"
dv_date.error = "Insira uma data válida (DD/MM/AAAA) ou deixe em branco. Texto não é permitido nesta coluna."
dv_date.sqref = f"F{first}:F{last_row}"
ws.add_data_validation(dv_date)
print(f"Validação de data adicionada em F{first}:F{last_row}")

wb.save(PATH)
print(f"\n✅ Salvo: {PATH}")

# ── Verificação ───────────────────────────────────────────────────────────────
wb2 = load_workbook(PATH)
ws2 = wb2['DB_ATIVIDADES']
print("\n=== Verificação pós-fix ===")
for r in range(4, 9):
    f = ws2.cell(row=r, column=6)
    g = ws2.cell(row=r, column=7)
    print(f"  L{r}: F val={repr(f.value)} fmt={f.number_format} | G fmt={g.number_format}")
print("\nCF ranges agora:")
for rng in ws2.conditional_formatting:
    print(f"  {rng.sqref}  ({len(list(ws2.conditional_formatting[rng]))} regras)")
